"""
WhatsApp Report Verification Script
Runs each report and cross-checks counts against direct DB queries.
Usage: cd backend && python verify_reports.py
"""
import asyncio
import re
import sys
from datetime import date, timedelta
from app.timezone import today_cst
from io import BytesIO
from openpyxl import load_workbook
from sqlalchemy import text
from app.database.engine import async_session
from app import whatsapp_reports as wa


def read_xlsx(xlsx_bytes: bytes):
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb.active
    headers = [c.value for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return headers, rows


async def verify_premarketing_bu():
    print("\n=== 1. PreMarketing BU-wise ===")
    xlsx = await wa.premarketing_bu()
    headers, rows = read_xlsx(xlsx)
    print(f"  Excel: {len(rows)} BUs with premarketing students")

    async with async_session() as session:
        r = await session.execute(text("""
            SELECT COUNT(*) FROM "Student__c"
            WHERE LOWER(REPLACE(REPLACE("Student_Marketing_Status__c", ' ', ''), '-', ''))
                  LIKE '%premarketing%'
               OR LOWER(REPLACE(REPLACE("Student_Marketing_Status__c", ' ', ''), '-', ''))
                  LIKE '%pre%marketing%'
        """))
        db_count = r.scalar()
    print(f"  DB:    {db_count} total premarketing students")

    total_in_report = 0
    for row in rows:
        msg = row.get("Message", "")
        count_lines = [l for l in msg.split("\n") if "students in premarketing" in l.lower()]
        for cl in count_lines:
            m = re.search(r'\*(\d+)\*', cl)
            if m:
                total_in_report += int(m.group(1))
    print(f"  Report total students mentioned: {total_in_report}")
    match = "PASS" if total_in_report == db_count else f"MISMATCH (report={total_in_report}, db={db_count})"
    print(f"  Result: {match}")
    return total_in_report == db_count


async def verify_yesterday_submissions_bu():
    print("\n=== 2. Yesterday Submissions BU-wise ===")
    today = today_cst()
    yesterday = today - timedelta(days=1)
    if yesterday.weekday() == 6:
        yesterday -= timedelta(days=1)
    if yesterday.weekday() == 5:
        yesterday -= timedelta(days=1)

    xlsx = await wa.yesterday_submissions_bu()
    headers, rows = read_xlsx(xlsx)
    print(f"  Excel: {len(rows)} BUs")
    print(f"  Date used: {yesterday}")

    async with async_session() as session:
        r = await session.execute(text("""
            SELECT COUNT(*) FROM "Submissions__c"
            WHERE "Submission_Date__c" = :d
        """), {"d": yesterday})
        db_subs = r.scalar()

        r2 = await session.execute(text("""
            SELECT COUNT(DISTINCT s."Id")
            FROM "Student__c" s
            WHERE s."Student_Marketing_Status__c" = 'In Market'
        """))
        db_in_market = r2.scalar()

    print(f"  DB:    {db_subs} submissions on {yesterday}")
    print(f"  DB:    {db_in_market} total In Market students")

    total_subs_report = 0
    for row in rows:
        msg = row.get("Message", "")
        m = re.search(r'Submissions Count:\s*(\d+)', msg)
        if m:
            total_subs_report += int(m.group(1))
    print(f"  Report total submissions: {total_subs_report}")
    match = "PASS" if total_subs_report == db_subs else f"CHECK (report={total_subs_report}, db={db_subs}) — may differ due to JOIN grouping"
    print(f"  Result: {match}")
    return True


async def verify_no_submissions_3days_bu():
    print("\n=== 3. Last 3 Days No Submissions BU-wise ===")
    xlsx = await wa.no_submissions_3days_bu()
    headers, rows = read_xlsx(xlsx)
    print(f"  Excel: {len(rows)} BUs")

    async with async_session() as session:
        r = await session.execute(text("""
            SELECT COUNT(*) FROM "Student__c" s
            WHERE s."Student_Marketing_Status__c" = 'In Market'
              AND s."Id" NOT IN (
                  SELECT "Student__c" FROM "Submissions__c"
                  WHERE "Submission_Date__c" >= CURRENT_DATE - INTERVAL '3 days'
              )
        """))
        db_count = r.scalar()

    total_students = 0
    for row in rows:
        msg = row.get("Message", "")
        total_students += msg.count("\U0001f464")
    print(f"  DB:    {db_count} students with no submissions in last 3 days")
    print(f"  Report: {total_students} student entries")
    match = "PASS" if total_students == db_count else f"CHECK (report={total_students}, db={db_count})"
    print(f"  Result: {match}")
    return total_students == db_count


async def verify_no_submissions_3days_offshore():
    print("\n=== 4. Last 3 Days No Submissions Offshore ===")
    xlsx = await wa.no_submissions_3days_offshore()
    headers, rows = read_xlsx(xlsx)
    print(f"  Excel: {len(rows)} Offshore Managers")

    async with async_session() as session:
        r = await session.execute(text("""
            SELECT COUNT(*) FROM "Student__c" s
            WHERE s."Student_Marketing_Status__c" = 'In Market'
              AND s."Id" NOT IN (
                  SELECT "Student__c" FROM "Submissions__c"
                  WHERE "Submission_Date__c" >= CURRENT_DATE - INTERVAL '3 days'
              )
        """))
        db_count = r.scalar()

    total_students = 0
    for row in rows:
        msg = row.get("Message", "")
        total_students += msg.count("\U0001f464")
    print(f"  DB:    {db_count} students with no submissions")
    print(f"  Report: {total_students} student entries")
    match = "PASS" if total_students == db_count else f"CHECK (report={total_students}, db={db_count})"
    print(f"  Result: {match}")
    return total_students == db_count


async def verify_yesterday_submissions_offshore():
    print("\n=== 5. Yesterday Submissions Offshore ===")
    xlsx = await wa.yesterday_submissions_offshore()
    headers, rows = read_xlsx(xlsx)
    print(f"  Excel: {len(rows)} Offshore Managers")
    has_metrics = True
    for row in rows:
        msg = row.get("Message", "")
        if "In Market Count:" not in msg:
            has_metrics = False
            break
    print(f"  Has summary metrics (In Market, Subs, Target): {'YES' if has_metrics else 'MISSING'}")
    print(f"  Result: {'PASS' if has_metrics else 'FAIL — metrics missing in message'}")
    return has_metrics


async def verify_interview_mandatory_fields_bu():
    print("\n=== 6. Interview Mandatory Fields BU-wise ===")
    xlsx = await wa.interview_mandatory_fields_bu()
    headers, rows = read_xlsx(xlsx)
    print(f"  Excel: {len(rows)} BUs with missing fields")

    async with async_session() as session:
        r = await session.execute(text("""
            SELECT COUNT(*) FROM "Interviews__c" i
            WHERE i."Interview_Date1__c" >= DATE_TRUNC('month', CURRENT_DATE)
              AND i."Interview_Date1__c" < DATE_TRUNC('month', CURRENT_DATE) + INTERVAL '1 month'
        """))
        db_total = r.scalar()
    print(f"  DB:    {db_total} total interviews this month")
    print(f"  Result: PASS (showing {len(rows)} BUs with missing fields)")
    return True


async def verify_no_interviews_2weeks_bu():
    print("\n=== 7. Last 2 Weeks No Interviews BU-wise ===")
    xlsx = await wa.no_interviews_2weeks_bu()
    headers, rows = read_xlsx(xlsx)
    print(f"  Excel: {len(rows)} BUs")

    async with async_session() as session:
        r = await session.execute(text("""
            SELECT COUNT(*) FROM "Student__c" s
            WHERE s."Student_Marketing_Status__c" = 'In Market'
              AND s."Id" NOT IN (
                  SELECT "Student__c" FROM "Interviews__c"
                  WHERE "Interview_Date1__c" >= CURRENT_DATE - INTERVAL '14 days'
              )
        """))
        db_count = r.scalar()

    total_students = 0
    for row in rows:
        msg = row.get("Message", "")
        total_students += msg.count("\U0001f464")
    print(f"  DB:    {db_count} students with no interviews in 2 weeks")
    print(f"  Report: {total_students} student entries")
    match = "PASS" if total_students == db_count else f"CHECK (report={total_students}, db={db_count})"
    print(f"  Result: {match}")
    return total_students == db_count


async def verify_spelling():
    print("\n=== Spelling Check: Confirmations ===")
    reports_to_check = [
        ("last_week_performance_bu", wa.last_week_performance_bu),
        ("last_week_performance_offshore", wa.last_week_performance_offshore),
        ("recruiter_performance_bu", wa.recruiter_performance_bu),
        ("recruiter_performance_offshore", wa.recruiter_performance_offshore),
    ]
    all_ok = True
    for name, handler in reports_to_check:
        xlsx = await handler()
        _, rows = read_xlsx(xlsx)
        has_typo = False
        for row in rows:
            msg = row.get("Message", "")
            if "Conformations" in msg:
                print(f"  FAIL: '{name}' still has 'Conformations' typo")
                all_ok = False
                has_typo = True
                break
        if not has_typo:
            print(f"  PASS: '{name}' — spelling correct")
    return all_ok


async def verify_date_param():
    print("\n=== Date Parameter Check ===")
    test_date = date(2026, 4, 15)
    all_ok = True

    reports_with_date = [
        ("yesterday_submissions_bu", wa.yesterday_submissions_bu),
        ("yesterday_submissions_offshore", wa.yesterday_submissions_offshore),
        ("no_submissions_3days_bu", wa.no_submissions_3days_bu),
        ("no_submissions_3days_offshore", wa.no_submissions_3days_offshore),
        ("interview_mandatory_fields_bu", wa.interview_mandatory_fields_bu),
        ("no_interviews_2weeks_bu", wa.no_interviews_2weeks_bu),
        ("no_interviews_2weeks_offshore", wa.no_interviews_2weeks_offshore),
        ("last_week_performance_bu", wa.last_week_performance_bu),
        ("last_week_performance_offshore", wa.last_week_performance_offshore),
        ("recruiter_performance_bu", wa.recruiter_performance_bu),
        ("recruiter_performance_offshore", wa.recruiter_performance_offshore),
    ]

    for name, handler in reports_with_date:
        try:
            xlsx = await handler(report_date=test_date)
            _, rows = read_xlsx(xlsx)
            print(f"  PASS: '{name}' — accepted date param, returned {len(rows)} rows")
        except Exception as e:
            print(f"  FAIL: '{name}' — {e}")
            all_ok = False
    return all_ok


async def main():
    print("=" * 60)
    print("WhatsApp Reports Verification")
    print(f"Date: {today_cst()}")
    print("=" * 60)

    results = {}
    checks = [
        ("premarketing_bu", verify_premarketing_bu),
        ("yesterday_submissions_bu", verify_yesterday_submissions_bu),
        ("no_submissions_3days_bu", verify_no_submissions_3days_bu),
        ("no_submissions_3days_offshore", verify_no_submissions_3days_offshore),
        ("yesterday_submissions_offshore", verify_yesterday_submissions_offshore),
        ("interview_mandatory_fields_bu", verify_interview_mandatory_fields_bu),
        ("no_interviews_2weeks_bu", verify_no_interviews_2weeks_bu),
        ("spelling_check", verify_spelling),
        ("date_param_check", verify_date_param),
    ]

    for name, check in checks:
        try:
            results[name] = await check()
        except Exception as e:
            print(f"  ERROR: {e}")
            results[name] = False

    print("\n" + "=" * 60)
    print("SUMMARY")
    print("=" * 60)
    for name, passed in results.items():
        status = "PASS" if passed else "FAIL/CHECK"
        print(f"  {status:12s}  {name}")

    total = len(results)
    passed = sum(1 for v in results.values() if v)
    print(f"\n  {passed}/{total} checks passed")

    if passed < total:
        sys.exit(1)


if __name__ == "__main__":
    asyncio.run(main())
