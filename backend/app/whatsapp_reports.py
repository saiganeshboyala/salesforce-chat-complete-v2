"""
WhatsApp Report Generator — generates BU-wise / Manager-wise
formatted WhatsApp messages from the PostgreSQL database and
returns them as Excel files (BU Name | Message).
"""
import io
import logging
from datetime import datetime, date, timedelta
from sqlalchemy import text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from app.database.engine import async_session

logger = logging.getLogger(__name__)


def _xlsx_bytes(rows: list[dict], columns: list[str], sheet_name: str = "Report") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    hdr_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    hdr_fill = PatternFill(start_color="2F5486", end_color="2F5486", fill_type="solid")

    for ci, h in enumerate(columns, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="left", vertical="center")

    for ri, row in enumerate(rows, 2):
        for ci, h in enumerate(columns, 1):
            cell = ws.cell(row=ri, column=ci, value=row.get(h, ""))
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    for ci, h in enumerate(columns, 1):
        if h.lower() == "message":
            ws.column_dimensions[get_column_letter(ci)].width = 120
        else:
            ws.column_dimensions[get_column_letter(ci)].width = 30

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ────────────────────────────────────────────────────────────
# 1. PreMarketing BU-wise
# ────────────────────────────────────────────────────────────

async def premarketing_bu() -> bytes:
    sql = text("""
        SELECT
            COALESCE(m."Name", 'Unknown') AS bu_name,
            s."Batch__c" AS batch,
            s."Name" AS student_name
        FROM "Student__c" s
        LEFT JOIN "Manager__c" m ON s."Manager__c" = m."Id"
        WHERE LOWER(REPLACE(REPLACE(s."Student_Marketing_Status__c", ' ', ''), '-', ''))
              LIKE '%premarketing%'
           OR LOWER(REPLACE(REPLACE(s."Student_Marketing_Status__c", ' ', ''), '-', ''))
              LIKE '%pre%marketing%'
        ORDER BY m."Name", s."Batch__c", s."Name"
    """)

    async with async_session() as session:
        result = await session.execute(sql)
        rows = result.fetchall()

    bu_data = {}
    for bu_name, batch, student_name in rows:
        bu = bu_name or "Unknown"
        b = batch if batch and str(batch).strip() and str(batch).strip().upper() != "NA" else None
        if bu not in bu_data:
            bu_data[bu] = {}
        if b:
            if b not in bu_data[bu]:
                bu_data[bu][b] = []
            bu_data[bu][b].append(student_name or "Unknown")

    output_rows = []
    for bu, batches in sorted(bu_data.items()):
        total = sum(len(students) for students in batches.values())
        lines = []
        lines.append("\U0001f6a8 Premarketing\n")
        lines.append(f"\U0001f3e2 *BU: {bu}*")
        lines.append(f"Still I can see *{total}* students in premarketing.")
        lines.append("This delay is not acceptable.\n")
        lines.append("Confirm the timeline for moving each student to *In Market* status.")
        lines.append("Individual acknowledgments must be provided for every student update without exception.")
        lines.append("━" * 20)

        for batch_name, students in sorted(batches.items()):
            lines.append(f"\U0001f4c5 *{batch_name}*")
            lines.append(f"\U0001f4ca Total Premarketing Students: {len(students)}")
            for s in students:
                lines.append(f"\U0001f464 {s}")
            lines.append("━" * 20)

        lines.append("⚠️ *Action Required*")
        lines.append("\U0001f449 Confirm movement count by *EOD today without fail*")

        output_rows.append({"BU Name": bu, "Message": "\n".join(lines)})

    return _xlsx_bytes(output_rows, ["BU Name", "Message"], "Premarketing BU")


# ────────────────────────────────────────────────────────────
# 2. Yesterday Submission Report BU-wise
# ────────────────────────────────────────────────────────────

async def yesterday_submissions_bu() -> bytes:
    today = date.today()
    yesterday = today - timedelta(days=1)
    if yesterday.weekday() == 6:
        yesterday = yesterday - timedelta(days=1)
    if yesterday.weekday() == 5:
        yesterday = yesterday - timedelta(days=1)
    date_str = yesterday.strftime("%B %d, %Y")

    bu_summary_sql = text("""
        SELECT
            COALESCE(m."Name", 'Unknown') AS bu_name,
            COUNT(DISTINCT s."Id") FILTER (WHERE s."Student_Marketing_Status__c" = 'In Market') AS in_market,
            COUNT(sub."Id") AS sub_count
        FROM "Manager__c" m
        LEFT JOIN "Student__c" s ON s."Manager__c" = m."Id"
        LEFT JOIN "Submissions__c" sub ON sub."Student__c" = s."Id"
            AND sub."Submission_Date__c" = :yesterday
        GROUP BY m."Name"
        ORDER BY m."Name"
    """)

    detail_sql = text("""
        SELECT
            COALESCE(m."Name", 'Unknown') AS bu_name,
            s."Offshore_Manager_Name__c" AS offshore_mgr,
            s."Name" AS student_name,
            s."Recruiter_Name__c" AS recruiter,
            s."Last_Submission_Date__c" AS last_sub_date,
            sub."Id" AS sub_id
        FROM "Submissions__c" sub
        JOIN "Student__c" s ON sub."Student__c" = s."Id"
        LEFT JOIN "Manager__c" m ON s."Manager__c" = m."Id"
        WHERE sub."Submission_Date__c" = :yesterday
          AND sub."Id" IS NOT NULL
          AND s."Last_Submission_Date__c" IS NOT NULL
        ORDER BY m."Name", s."Offshore_Manager_Name__c", s."Name"
    """)

    async with async_session() as session:
        summary_res = await session.execute(bu_summary_sql, {"yesterday": yesterday})
        summary_rows = summary_res.fetchall()
        detail_res = await session.execute(detail_sql, {"yesterday": yesterday})
        detail_rows = detail_res.fetchall()

    bu_summary = {}
    for bu_name, in_market, sub_count in summary_rows:
        bu = bu_name or "Unknown"
        im = in_market or 0
        sc = sub_count or 0
        target = round((sc / (im * 2) * 100), 1) if im > 0 else 0.0
        bu_summary[bu] = {"in_market": im, "sub_count": sc, "target": target}

    bu_details = {}
    for bu_name, offshore_mgr, student_name, recruiter, last_sub_date, sub_id in detail_rows:
        bu = bu_name or "Unknown"
        if bu not in bu_details:
            bu_details[bu] = {}
        mgr = offshore_mgr or "NAN"
        if mgr not in bu_details[bu]:
            bu_details[bu][mgr] = []
        lsd = f"{last_sub_date.month}/{last_sub_date.day}/{last_sub_date.year}" if last_sub_date else ""
        if lsd:
            bu_details[bu][mgr].append({
                "student": student_name or "Unknown",
                "recruiter": recruiter or "NAN",
                "last_sub": lsd,
            })

    def _perf_line(target_pct):
        if target_pct >= 100:
            return "\n\U0001f44d Excellent execution. Keep up the momentum."
        if target_pct >= 70:
            return "\n\U0001f44d Good progress. Push for higher submissions."
        if target_pct >= 40:
            return "\n\U0001f449 Performance is below expectations. Strengthen daily submission activity."
        return (
            "\n\U0001f449 Performance is significantly below expectations. "
            "Immediate focus is required on increasing submissions and maintaining daily activity."
        )

    output_rows = []
    for bu in sorted(set(list(bu_summary.keys()) + list(bu_details.keys()))):
        s = bu_summary.get(bu, {"in_market": 0, "sub_count": 0, "target": 0.0})
        lines = []
        lines.append(f"\U0001f4ca Yesterday Submission Report | \U0001f4c5 {date_str}\n")
        lines.append(f"*{bu}*\n")
        lines.append(f"\U0001f465 In Market Count: {s['in_market']}")
        lines.append(f"\U0001f4e4 Submissions Count: {s['sub_count']}")
        lines.append(f"\U0001f3af %% Target: {s['target']}")

        details = bu_details.get(bu, {})
        for mgr, students in sorted(details.items()):
            lines.append(f"\nOffshore Manager Name: {mgr}")
            for st in students:
                lines.append(
                    f"\U0001f464 *{st['student']}* | \U0001f465 Recruiter: {st['recruiter']} "
                    f"| \U0001f4c5 Last Submission: {st['last_sub']}"
                )
            lines.append(_perf_line(s['target']))

        if not details:
            lines.append(_perf_line(s['target']))

        output_rows.append({"BU": bu, "Message": "\n".join(lines)})

    return _xlsx_bytes(output_rows, ["BU", "Message"], "Yesterday Submissions BU")


# ────────────────────────────────────────────────────────────
# 3. Last 3 Days No Submissions BU-wise
# ────────────────────────────────────────────────────────────

async def no_submissions_3days_bu() -> bytes:
    today = date.today()

    sql = text("""
        SELECT
            COALESCE(m."Name", 'Unknown') AS bu_name,
            s."Offshore_Manager_Name__c" AS offshore_mgr,
            s."Recruiter_Name__c" AS recruiter,
            s."Name" AS student_name,
            s."Last_Submission_Date__c" AS last_sub_date
        FROM "Student__c" s
        LEFT JOIN "Manager__c" m ON s."Manager__c" = m."Id"
        WHERE s."Student_Marketing_Status__c" = 'In Market'
          AND s."Id" NOT IN (
              SELECT "Student__c" FROM "Submissions__c"
              WHERE "Submission_Date__c" >= CURRENT_DATE - INTERVAL '3 days'
          )
        ORDER BY m."Name", s."Offshore_Manager_Name__c", s."Recruiter_Name__c", s."Name"
    """)

    async with async_session() as session:
        result = await session.execute(sql)
        rows = result.fetchall()

    bu_data = {}
    for bu_name, offshore_mgr, recruiter, student_name, last_sub_date in rows:
        bu = bu_name or "Unknown"
        mgr = offshore_mgr or "NAN"
        rec = recruiter or "NAN"
        if bu not in bu_data:
            bu_data[bu] = {}
        if mgr not in bu_data[bu]:
            bu_data[bu][mgr] = {}
        if rec not in bu_data[bu][mgr]:
            bu_data[bu][mgr][rec] = []

        if last_sub_date:
            delta = (today - last_sub_date).days
            gap = f"{delta} days"
        else:
            gap = "Till Date No Submissions"

        bu_data[bu][mgr][rec].append({"name": student_name or "Unknown", "gap": gap, "days": (today - last_sub_date).days if last_sub_date else 9999})

    output_rows = []
    for bu, managers in sorted(bu_data.items()):
        lines = []
        lines.append("\U0001f4ca Last 3 Days No Submissions Report\n")
        lines.append(f"*{bu}*")

        for mgr, recruiters in sorted(managers.items()):
            lines.append(f"\n\U0001f9d1‍\U0001f4bc *Offshore Manager: {mgr}*")
            all_days = []
            for rec, students in sorted(recruiters.items()):
                lines.append(f"\U0001f465 Recruiter: {rec}")
                for st in students:
                    lines.append(f"\U0001f464 *{st['name']}* | ⏳ No Submissions: {st['gap']}")
                    if st['days'] < 9999:
                        all_days.append(st['days'])

            if not all_days:
                action = "\U0001f6a8 Critical: Immediate escalation required. No activity observed across team."
            elif all(d > 10 for d in all_days):
                action = "\U0001f6a8 Critical: Immediate escalation required. No activity observed across team."
            elif sum(1 for d in all_days if d > 7) / len(all_days) > 0.7:
                action = "⚠️ High Concern: Strong push required. Enforce strict daily submission targets."
            elif any(d > 3 for d in all_days):
                action = "⚡ Moderate: Needs improvement. Ensure consistent daily tracking."
            else:
                action = "✅ Stable: Maintain consistency and improve conversions."

            lines.append(f"\nAction:\n{action}")
            lines.append("━" * 20)

        output_rows.append({"BU Name": bu, "Message": "\n".join(lines)})

    return _xlsx_bytes(output_rows, ["BU Name", "Message"], "No Submissions 3 Days BU")


# ────────────────────────────────────────────────────────────
# 4. Last 3 Days No Submissions Offshore Manager-wise
# ────────────────────────────────────────────────────────────

async def no_submissions_3days_offshore() -> bytes:
    today = date.today()
    date_str = today.strftime("%B %d, %Y")

    sql = text("""
        SELECT
            s."Offshore_Manager_Name__c" AS offshore_mgr,
            s."Recruiter_Name__c" AS recruiter,
            s."Name" AS student_name,
            s."Last_Submission_Date__c" AS last_sub_date
        FROM "Student__c" s
        WHERE s."Student_Marketing_Status__c" = 'In Market'
          AND s."Id" NOT IN (
              SELECT "Student__c" FROM "Submissions__c"
              WHERE "Submission_Date__c" >= CURRENT_DATE - INTERVAL '3 days'
          )
        ORDER BY s."Offshore_Manager_Name__c", s."Recruiter_Name__c", s."Name"
    """)

    async with async_session() as session:
        result = await session.execute(sql)
        rows = result.fetchall()

    mgr_data = {}
    for offshore_mgr, recruiter, student_name, last_sub_date in rows:
        mgr = offshore_mgr or "NAN"
        rec = recruiter or "NAN"
        if mgr not in mgr_data:
            mgr_data[mgr] = {}
        if rec not in mgr_data[mgr]:
            mgr_data[mgr][rec] = []

        if last_sub_date:
            delta = (today - last_sub_date).days
            gap = f"{delta} days"
        else:
            gap = "till today no submissions"

        mgr_data[mgr][rec].append({"name": student_name or "Unknown", "gap": gap, "days": (today - last_sub_date).days if last_sub_date else 9999})

    output_rows = []
    for mgr, recruiters in sorted(mgr_data.items()):
        lines = []
        lines.append(f"\U0001f4ca Last 3 Days No Submissions Report\n")
        lines.append(f"*{mgr}*")

        all_days = []
        for rec, students in sorted(recruiters.items()):
            lines.append(f"\n\U0001f465 Recruiter: {rec}")
            for st in students:
                lines.append(f"\U0001f464 *{st['name']}* | ⏳ No Submissions: {st['gap']}")
                if st['days'] < 9999:
                    all_days.append(st['days'])

        if not all_days:
            action = "\U0001f6a8 Critical: No submissions data available. Investigate immediately."
        elif all(d > 10 for d in all_days):
            action = "\U0001f6a8 Critical: Immediate escalation required. No activity observed across team."
        elif sum(1 for d in all_days if d > 7) / len(all_days) > 0.7:
            action = "⚠️ High Concern: Strong push required. Enforce strict daily submission targets."
        elif any(3 < d <= 7 for d in all_days):
            action = "⚡ Moderate: Needs improvement. Ensure consistent daily tracking."
        else:
            action = "\U0001f4c8 Push for consistency. Maintain submission discipline, ensure regular vendor engagement, and avoid gaps in activity."

        lines.append(f"\n⚡ Action:\n{action}")

        output_rows.append({"Offshore Manager": mgr, "Message": "\n".join(lines)})

    return _xlsx_bytes(output_rows, ["Offshore Manager", "Message"], "No Submissions 3 Days OM")


# ────────────────────────────────────────────────────────────
# 5. Yesterday Submission Report Offshore Manager-wise
# ────────────────────────────────────────────────────────────

async def yesterday_submissions_offshore() -> bytes:
    today = date.today()
    yesterday = today - timedelta(days=1)
    if yesterday.weekday() == 6:
        yesterday -= timedelta(days=1)
    if yesterday.weekday() == 5:
        yesterday -= timedelta(days=1)
    date_str = yesterday.strftime("%B %d, %Y")

    summary_sql = text("""
        SELECT
            s."Offshore_Manager_Name__c" AS offshore_mgr,
            COUNT(DISTINCT s."Id") FILTER (WHERE s."Student_Marketing_Status__c" = 'In Market') AS in_market,
            COUNT(sub."Id") AS sub_count
        FROM "Student__c" s
        LEFT JOIN "Submissions__c" sub ON sub."Student__c" = s."Id"
            AND sub."Submission_Date__c" = :yesterday
        WHERE s."Offshore_Manager_Name__c" IS NOT NULL
        GROUP BY s."Offshore_Manager_Name__c"
        ORDER BY s."Offshore_Manager_Name__c"
    """)

    detail_sql = text("""
        SELECT
            s."Offshore_Manager_Name__c" AS offshore_mgr,
            s."Name" AS student_name,
            s."Recruiter_Name__c" AS recruiter,
            s."Last_Submission_Date__c" AS last_sub_date
        FROM "Submissions__c" sub
        JOIN "Student__c" s ON sub."Student__c" = s."Id"
        WHERE sub."Submission_Date__c" = :yesterday
        ORDER BY s."Offshore_Manager_Name__c", s."Name"
    """)

    async with async_session() as session:
        summary_res = await session.execute(summary_sql, {"yesterday": yesterday})
        summary_rows = summary_res.fetchall()
        detail_res = await session.execute(detail_sql, {"yesterday": yesterday})
        detail_rows = detail_res.fetchall()

    mgr_summary = {}
    for offshore_mgr, in_market, sub_count in summary_rows:
        mgr = offshore_mgr or "NAN"
        im = in_market or 0
        sc = sub_count or 0
        target = round((sc / (im * 2) * 100), 1) if im > 0 else 0.0
        mgr_summary[mgr] = {"in_market": im, "sub_count": sc, "target": target}

    mgr_details = {}
    for offshore_mgr, student_name, recruiter, last_sub_date in detail_rows:
        mgr = offshore_mgr or "NAN"
        if mgr not in mgr_details:
            mgr_details[mgr] = []
        lsd = f"{last_sub_date.month}/{last_sub_date.day}/{last_sub_date.year}" if last_sub_date else ""
        mgr_details[mgr].append({
            "student": student_name or "Unknown",
            "recruiter": recruiter or "NAN",
            "last_sub": lsd,
        })

    output_rows = []
    for mgr in sorted(set(list(mgr_summary.keys()) + list(mgr_details.keys()))):
        s = mgr_summary.get(mgr, {"in_market": 0, "sub_count": 0, "target": 0.0})
        lines = []
        lines.append(f"\U0001f4ca Yesterday Submission Report | \U0001f4c5 {date_str}\n")
        lines.append(f"*{mgr}*")

        details = mgr_details.get(mgr, [])
        for st in details:
            lines.append(
                f"\U0001f464 *{st['student']}* | \U0001f465 Recruiter: {st['recruiter']} "
                f"| \U0001f4c5 Last Submission: {st['last_sub']}"
            )

        if s['target'] >= 100:
            lines.append("\n\U0001f44d Excellent execution. Keep up the momentum.")
        elif s['target'] >= 70:
            lines.append("\n\U0001f44d Good progress. Push for higher submissions.")
        elif s['target'] >= 40:
            lines.append("\n\U0001f449 Performance is below expectations. Strengthen daily submission activity and improve planning to close the gap.")
        else:
            lines.append("\n\U0001f449 Performance is significantly below expectations. Immediate focus is required on increasing submissions and maintaining daily activity.")

        output_rows.append({"Offshore Manager Name": mgr, "Message": "\n".join(lines)})

    return _xlsx_bytes(output_rows, ["Offshore Manager Name", "Message"], "Yesterday Submissions OM")


# ────────────────────────────────────────────────────────────
# 6. Interview Mandatory Fields BU-wise
# ────────────────────────────────────────────────────────────

async def interview_mandatory_fields_bu() -> bytes:
    mandatory_fields = [
        ("Student_Otter_Performance__c", "Student Otter Performance"),
        ("Student_Technical_Explanation_Skill__c", "Student Technical Explanation Skill"),
        ("Any_Technical_Issues__c", "Any Technical Issues?"),
        ("Proxy_General_Issues__c", "Proxy General Issues"),
        ("Interview_Q_A__c", "Interview Q&A"),
        ("TechSupportName__c", "Tech Support: Name"),
        ("Final_Status__c", "Interview Final Status"),
    ]

    field_cols = ", ".join(f'i."{f[0]}"' for f in mandatory_fields)
    sql = text(f"""
        SELECT
            COALESCE(m."Name", 'Unknown') AS bu_name,
            s."Name" AS student_name,
            {field_cols},
            i."Id" AS interview_id
        FROM "Interviews__c" i
        JOIN "Student__c" s ON i."Student__c" = s."Id"
        LEFT JOIN "Manager__c" m ON s."Manager__c" = m."Id"
        WHERE i."Interview_Date1__c" >= DATE_TRUNC('month', CURRENT_DATE)
          AND i."Interview_Date1__c" < DATE_TRUNC('month', CURRENT_DATE) + INTERVAL '1 month'
        ORDER BY m."Name", s."Name"
    """)

    async with async_session() as session:
        result = await session.execute(sql)
        rows = result.fetchall()

    bu_data = {}
    for row in rows:
        bu = row[0] or "Unknown"
        student = row[1] or "Unknown"
        field_values = row[2:2 + len(mandatory_fields)]
        interview_id = row[-1]

        missing = []
        for i, (field_name, label) in enumerate(mandatory_fields):
            val = field_values[i]
            if val is None or str(val).strip() == "":
                missing.append(label)

        if not missing:
            continue

        if bu not in bu_data:
            bu_data[bu] = {"students": {}, "interview_count": 0}
        bu_data[bu]["interview_count"] += 1
        if student not in bu_data[bu]["students"]:
            bu_data[bu]["students"][student] = []
        bu_data[bu]["students"][student].extend(missing)

    output_rows = []
    for bu, data in sorted(bu_data.items()):
        lines = []
        lines.append(
            "⚠️ u r getting this msg because u r not updating the data points in interviews, "
            "I am following on this every day still i can see same delay count of interviews, "
            "Please Update everyday before you sleep.\n"
        )
        lines.append(f"\U0001f4ca Interview Count: {data['interview_count']}\n")
        lines.append("\U0001f6a8 Action Required: Below students have missing fields.\n")

        for student, missing_fields in sorted(data["students"].items()):
            unique_missing = list(dict.fromkeys(missing_fields))
            lines.append(f"\U0001f464 *{student}*")
            for field in unique_missing:
                lines.append(f" • {field} → *Fill this data asap*")

        output_rows.append({"BU Name": bu, "Message": "\n".join(lines)})

    return _xlsx_bytes(output_rows, ["BU Name", "Message"], "Interview Fields BU")


# ────────────────────────────────────────────────────────────
# 7. Last 2 Weeks No Interviews BU-wise
# ────────────────────────────────────────────────────────────

async def no_interviews_2weeks_bu() -> bytes:
    today = date.today()

    sql = text("""
        SELECT
            COALESCE(m."Name", 'Unknown') AS bu_name,
            s."Offshore_Manager_Name__c" AS offshore_mgr,
            s."Recruiter_Name__c" AS recruiter,
            s."Name" AS student_name,
            s."Days_in_Market_Business__c" AS days_in_market,
            s."Recent_Past_Interview_Date__c" AS recent_interview
        FROM "Student__c" s
        LEFT JOIN "Manager__c" m ON s."Manager__c" = m."Id"
        WHERE s."Student_Marketing_Status__c" = 'In Market'
          AND s."Id" NOT IN (
              SELECT "Student__c" FROM "Interviews__c"
              WHERE "Interview_Date1__c" >= CURRENT_DATE - INTERVAL '14 days'
          )
        ORDER BY m."Name", s."Offshore_Manager_Name__c", s."Recruiter_Name__c", s."Name"
    """)

    async with async_session() as session:
        result = await session.execute(sql)
        rows = result.fetchall()

    bu_data = {}
    for bu_name, offshore_mgr, recruiter, student_name, days_in_market, recent_interview in rows:
        bu = bu_name or "Unknown"
        mgr = offshore_mgr or "NAN"
        rec = recruiter or "NAN"

        if recent_interview:
            since_days = (today - recent_interview).days
        elif days_in_market:
            since_days = int(days_in_market)
        else:
            since_days = None

        if bu not in bu_data:
            bu_data[bu] = {}
        if mgr not in bu_data[bu]:
            bu_data[bu][mgr] = []

        bu_data[bu][mgr].append({
            "student": student_name or "Unknown",
            "recruiter": rec,
            "since_days": since_days,
        })

    output_rows = []
    for bu, managers in sorted(bu_data.items()):
        lines = []
        lines.append("\U0001f4ca Last 2 Weeks – No Interviews for these students\n")
        lines.append(f"*{bu}*")

        for mgr, students in sorted(managers.items()):
            lines.append(f"\n\U0001f9d1‍\U0001f4bc *Offshore Manager: {mgr}*")

            for st in students:
                days_str = f"*{st['since_days']}*" if st['since_days'] is not None else "NA"
                lines.append(
                    f"\U0001f464 *{st['student']}* | \U0001f465 Recruiter: {st['recruiter']} "
                    f"| ⏳ since {days_str} days no interview"
                )

            valid_days = [st['since_days'] for st in students if st['since_days'] is not None]
            if not valid_days:
                action = "⚠️ Data Gap: Update interview dates and days in market for accurate tracking."
            else:
                avg = sum(valid_days) / len(valid_days)
                if avg > 60:
                    action = "\U0001f6a8 Critical: No interviews for extended period. Immediate escalation required. Conduct daily vendor connects, resume revamp, and enforce strict submission targets."
                elif avg > 30:
                    action = "⚠️ High Priority: Significant interview gap. Increase submissions quality, prioritize vendor follow-ups, and ensure profile alignment immediately."
                elif avg > 14:
                    action = "⚡ Moderate: Interview activity needs improvement. Focus on targeted submissions and strengthen client engagement."
                else:
                    action = "✅ Stable: Maintain consistency and push for interview conversions."

            lines.append(f"\n⚡ *Action:*\n{action}")

        output_rows.append({"BU Name": bu, "Message": "\n".join(lines)})

    return _xlsx_bytes(output_rows, ["BU Name", "Message"], "No Interviews 2 Weeks BU")


# ────────────────────────────────────────────────────────────
# 8. Last 2 Weeks No Interviews Offshore Manager-wise
# ────────────────────────────────────────────────────────────

async def no_interviews_2weeks_offshore() -> bytes:
    today = date.today()

    sql = text("""
        SELECT
            s."Offshore_Manager_Name__c" AS offshore_mgr,
            s."Recruiter_Name__c" AS recruiter,
            s."Name" AS student_name,
            s."Days_in_Market_Business__c" AS days_in_market,
            s."Recent_Past_Interview_Date__c" AS recent_interview
        FROM "Student__c" s
        WHERE s."Student_Marketing_Status__c" = 'In Market'
          AND s."Id" NOT IN (
              SELECT "Student__c" FROM "Interviews__c"
              WHERE "Interview_Date1__c" >= CURRENT_DATE - INTERVAL '14 days'
          )
        ORDER BY s."Offshore_Manager_Name__c", s."Recruiter_Name__c", s."Name"
    """)

    async with async_session() as session:
        result = await session.execute(sql)
        rows = result.fetchall()

    mgr_data = {}
    for offshore_mgr, recruiter, student_name, days_in_market, recent_interview in rows:
        mgr = offshore_mgr or "NAN"
        rec = recruiter or "NAN"

        if recent_interview:
            since_days = (today - recent_interview).days
        elif days_in_market:
            since_days = int(days_in_market)
        else:
            since_days = None

        if mgr not in mgr_data:
            mgr_data[mgr] = []

        mgr_data[mgr].append({
            "student": student_name or "Unknown",
            "recruiter": rec,
            "since_days": since_days,
        })

    output_rows = []
    for mgr, students in sorted(mgr_data.items()):
        lines = []
        lines.append("\U0001f4ca Last 2 Weeks No Interviews Report\n")
        lines.append(f"*{mgr}*")

        for st in students:
            days_str = f"*{st['since_days']} days*" if st['since_days'] is not None else "NA"
            lines.append(
                f"\U0001f464 *{st['student']}* | \U0001f465 Recruiter: {st['recruiter']} "
                f"| ⏳ No Interviews: {days_str}"
            )

        valid_days = [st['since_days'] for st in students if st['since_days'] is not None]
        if not valid_days:
            action = "⚠️ Data Gap: Update interview dates for accurate tracking."
        else:
            avg = sum(valid_days) / len(valid_days)
            if avg > 60:
                action = "\U0001f6a8 Critical: No interviews for extended period. Immediate escalation required. Conduct daily vendor connects, resume revamp, and enforce strict submission targets."
            elif avg > 30:
                action = "⚠️ High Priority: Significant interview gap. Increase submissions quality, prioritize vendor follow-ups, and ensure profile alignment immediately."
            elif avg > 14:
                action = "⚡ Moderate: Strengthen submissions quality, increase client follow-ups, and ensure profile alignment without delay."
            else:
                action = "✅ Stable: Maintain consistency and push for interview conversions."

        lines.append(f"\n⚠️ *Action Required:*\n\U0001f449 {action}")

        output_rows.append({"Offshore Manager Name": mgr, "Message": "\n".join(lines)})

    return _xlsx_bytes(output_rows, ["Offshore Manager Name", "Message"], "No Interviews 2 Weeks OM")


# ────────────────────────────────────────────────────────────
# 9. Last Week Submissions & Interviews BU-wise
# ────────────────────────────────────────────────────────────

async def last_week_performance_bu() -> bytes:
    sql = text("""
        SELECT
            COALESCE(m."Name", 'Unknown') AS bu_name,
            s."Offshore_Manager_Name__c" AS offshore_mgr,
            s."Recruiter_Name__c" AS recruiter,
            s."Name" AS student_name,
            COALESCE(s."Last_week_Submissions__c", 0) AS last_week_subs,
            COALESCE(s."Last_week_Interviews__c", 0) AS last_week_ints,
            CASE WHEN s."Verbal_Confirmation_Date__c" >= DATE_TRUNC('week', CURRENT_DATE) - INTERVAL '1 week'
                  AND s."Verbal_Confirmation_Date__c" < DATE_TRUNC('week', CURRENT_DATE)
                 THEN 1 ELSE 0 END AS confirmed
        FROM "Student__c" s
        LEFT JOIN "Manager__c" m ON s."Manager__c" = m."Id"
        WHERE s."Student_Marketing_Status__c" = 'In Market'
        ORDER BY m."Name", s."Offshore_Manager_Name__c", s."Recruiter_Name__c", s."Name"
    """)

    async with async_session() as session:
        result = await session.execute(sql)
        rows = result.fetchall()

    bu_data = {}
    for bu_name, offshore_mgr, recruiter, student_name, subs, ints, confirmed in rows:
        bu = bu_name or "Unknown"
        mgr = offshore_mgr or "NAN"
        rec = recruiter or "NAN"

        if bu not in bu_data:
            bu_data[bu] = {"managers": {}, "total_subs": 0, "total_ints": 0, "total_conf": 0, "in_market": 0}
        bu_data[bu]["total_subs"] += int(subs)
        bu_data[bu]["total_ints"] += int(ints)
        bu_data[bu]["total_conf"] += int(confirmed)
        bu_data[bu]["in_market"] += 1

        if mgr not in bu_data[bu]["managers"]:
            bu_data[bu]["managers"][mgr] = {}
        if rec not in bu_data[bu]["managers"][mgr]:
            bu_data[bu]["managers"][mgr][rec] = []

        bu_data[bu]["managers"][mgr][rec].append({
            "name": student_name or "Unknown",
            "subs": int(subs),
            "ints": int(ints),
            "conf": int(confirmed),
        })

    output_rows = []
    for bu, data in sorted(bu_data.items()):
        im = data["in_market"]
        target = round((data["total_subs"] / (im * 2 * 5) * 100), 2) if im > 0 else 0
        lines = []
        lines.append("\U0001f4ca Last Week Report\n")
        lines.append(f"*{bu}*")
        lines.append(f"\U0001f4e4 Submissions Count: {data['total_subs']}")
        lines.append(f"\U0001f3a4 Interview Count: {data['total_ints']}")
        lines.append(f"\U0001f465 In Market Count: {im}")
        lines.append(f"\U0001f3af %% Target: {target}")
        lines.append(f"✅ Conformations: {data['total_conf']}")

        for mgr, recruiters in sorted(data["managers"].items()):
            lines.append(f"\n\U0001f9d1‍\U0001f4bc Offshore Manager: *{mgr}*")
            for rec, students in sorted(recruiters.items()):
                lines.append(f"\U0001f465 Recruiter: *{rec}*")
                for st in students:
                    lines.append(
                        f"\U0001f464 Student: *{st['name']}* | \U0001f4e4 Submissions: {st['subs']} "
                        f"| \U0001f3a4 Interviews: {st['ints']} | ✅ Conformations: {st['conf']}"
                    )

        if data['total_subs'] == 0:
            action = "\U0001f6a8 Critical: No submissions. Immediate action required."
        elif data['total_ints'] == 0:
            action = "⚠️ No interviews. Improve quality and vendor follow-ups."
        elif data['total_conf'] == 0:
            action = "⚡ No closures. Focus on conversion and feedback."
        else:
            action = "✅ Good progress. Push for higher confirmations."

        lines.append(f"\n{action}")

        output_rows.append({"BU Name": bu, "Message": "\n".join(lines)})

    return _xlsx_bytes(output_rows, ["BU Name", "Message"], "Last Week Performance BU")


# ────────────────────────────────────────────────────────────
# 10. Last Week Submissions & Interviews Offshore Manager-wise
# ────────────────────────────────────────────────────────────

async def last_week_performance_offshore() -> bytes:
    sql = text("""
        SELECT
            s."Offshore_Manager_Name__c" AS offshore_mgr,
            s."Recruiter_Name__c" AS recruiter,
            s."Name" AS student_name,
            COALESCE(s."Last_week_Submissions__c", 0) AS last_week_subs,
            COALESCE(s."Last_week_Interviews__c", 0) AS last_week_ints,
            CASE WHEN s."Verbal_Confirmation_Date__c" >= DATE_TRUNC('week', CURRENT_DATE) - INTERVAL '1 week'
                  AND s."Verbal_Confirmation_Date__c" < DATE_TRUNC('week', CURRENT_DATE)
                 THEN 1 ELSE 0 END AS confirmed
        FROM "Student__c" s
        WHERE s."Student_Marketing_Status__c" = 'In Market'
        ORDER BY s."Offshore_Manager_Name__c", s."Recruiter_Name__c", s."Name"
    """)

    async with async_session() as session:
        result = await session.execute(sql)
        rows = result.fetchall()

    mgr_data = {}
    for offshore_mgr, recruiter, student_name, subs, ints, confirmed in rows:
        mgr = offshore_mgr or "NAN"
        rec = recruiter or "NAN"

        if mgr not in mgr_data:
            mgr_data[mgr] = {"recruiters": {}, "total_subs": 0, "total_ints": 0, "total_conf": 0, "in_market": 0}
        mgr_data[mgr]["total_subs"] += int(subs)
        mgr_data[mgr]["total_ints"] += int(ints)
        mgr_data[mgr]["total_conf"] += int(confirmed)
        mgr_data[mgr]["in_market"] += 1

        if rec not in mgr_data[mgr]["recruiters"]:
            mgr_data[mgr]["recruiters"][rec] = []

        mgr_data[mgr]["recruiters"][rec].append({
            "name": student_name or "Unknown",
            "subs": int(subs),
            "ints": int(ints),
            "conf": int(confirmed),
        })

    output_rows = []
    for mgr, data in sorted(mgr_data.items()):
        im = data["in_market"]
        target = round((data["total_subs"] / (im * 2 * 5) * 100), 2) if im > 0 else 0
        lines = []
        lines.append("\U0001f4ca Last Week Report\n")
        lines.append(f"*{mgr}*")
        lines.append(f"\U0001f4e4 Submissions Count: {data['total_subs']}")
        lines.append(f"\U0001f3a4 Interview Count: {data['total_ints']}")
        lines.append(f"\U0001f465 In Market Count: {im}")
        lines.append(f"\U0001f3af %% Target: {target}")
        lines.append(f"✅ Conformations: {data['total_conf']}")

        for rec, students in sorted(data["recruiters"].items()):
            lines.append(f"\n\U0001f465 Recruiter: *{rec}*")
            for st in students:
                lines.append(
                    f"\U0001f464 Student: *{st['name']}* | \U0001f4e4 Submissions: {st['subs']} "
                    f"| \U0001f3a4 Interviews: {st['ints']} | ✅ Conformations: {st['conf']}"
                )

        if data['total_subs'] == 0:
            action = "\U0001f6a8 Critical: No submissions. Immediate action required."
        elif data['total_ints'] == 0:
            action = "⚠️ No interviews. Improve quality and vendor follow-ups."
        elif data['total_conf'] == 0:
            action = "⚡ Improve conversions to offers."
        else:
            action = "✅ Good progress. Push for higher confirmations."

        lines.append(f"\n{action}")

        output_rows.append({"Offshore Manager Name": mgr, "Message": "\n".join(lines)})

    return _xlsx_bytes(output_rows, ["Offshore Manager Name", "Message"], "Last Week Performance OM")


# ────────────────────────────────────────────────────────────
# 11. Recruiter Last Week Performance BU-wise
# ────────────────────────────────────────────────────────────

async def recruiter_performance_bu() -> bytes:
    sql = text("""
        SELECT
            COALESCE(m."Name", 'Unknown') AS bu_name,
            s."Offshore_Manager_Name__c" AS offshore_mgr,
            s."Recruiter_Name__c" AS recruiter,
            s."Name" AS student_name,
            COALESCE(s."Last_week_Submissions__c", 0) AS last_week_subs,
            COALESCE(s."Last_week_Interviews__c", 0) AS last_week_ints,
            CASE WHEN s."Verbal_Confirmation_Date__c" >= DATE_TRUNC('week', CURRENT_DATE) - INTERVAL '1 week'
                  AND s."Verbal_Confirmation_Date__c" < DATE_TRUNC('week', CURRENT_DATE)
                 THEN 1 ELSE 0 END AS confirmed
        FROM "Student__c" s
        LEFT JOIN "Manager__c" m ON s."Manager__c" = m."Id"
        WHERE s."Student_Marketing_Status__c" = 'In Market'
        ORDER BY s."Recruiter_Name__c", s."Offshore_Manager_Name__c", m."Name", s."Name"
    """)

    async with async_session() as session:
        result = await session.execute(sql)
        rows = result.fetchall()

    rec_data = {}
    for bu_name, offshore_mgr, recruiter, student_name, subs, ints, confirmed in rows:
        bu = bu_name or "Unknown"
        mgr = offshore_mgr or "NAN"
        rec = recruiter or "NAN"
        key = (rec, mgr, bu)

        if key not in rec_data:
            rec_data[key] = {"students": [], "total_subs": 0, "total_ints": 0, "total_conf": 0, "in_market": 0}
        rec_data[key]["total_subs"] += int(subs)
        rec_data[key]["total_ints"] += int(ints)
        rec_data[key]["total_conf"] += int(confirmed)
        rec_data[key]["in_market"] += 1
        rec_data[key]["students"].append({
            "name": student_name or "Unknown",
            "subs": int(subs),
            "ints": int(ints),
            "conf": int(confirmed),
        })

    output_rows = []
    for (rec, mgr, bu), data in sorted(rec_data.items()):
        im = data["in_market"]
        lines = []
        lines.append(f"\U0001f4ca Last Week Recruiter Report\n")
        lines.append(f"*{rec}*")
        lines.append(f"\U0001f9d1‍\U0001f4bc Offshore Manager: *{mgr}*")
        lines.append(f"\U0001f3e2 BU: *{bu}*")
        lines.append(f"\U0001f4e4 Submissions: {data['total_subs']}")
        lines.append(f"\U0001f3a4 Interviews: {data['total_ints']}")
        lines.append(f"\U0001f465 In Market Count: {im}")
        lines.append(f"✅ Conformations: {data['total_conf']}")

        for st in data["students"]:
            lines.append(
                f"\n\U0001f464 Student: *{st['name']}* | \U0001f4e4 Submissions: {st['subs']} "
                f"| \U0001f3a4 Interviews: {st['ints']} | ✅ *Conformations:* {st['conf']}"
            )

        output_rows.append({
            "Recruiter": rec,
            "Offshore Manager": mgr,
            "BU": bu,
            "Message": "\n".join(lines),
        })

    return _xlsx_bytes(output_rows, ["Recruiter", "Offshore Manager", "BU", "Message"], "Recruiter Performance BU")


# ────────────────────────────────────────────────────────────
# 12. Recruiter Last Week Performance Offshore Manager-wise
# ────────────────────────────────────────────────────────────

async def recruiter_performance_offshore() -> bytes:
    sql = text("""
        SELECT
            s."Offshore_Manager_Name__c" AS offshore_mgr,
            COALESCE(m."Name", 'Unknown') AS bu_name,
            s."Recruiter_Name__c" AS recruiter,
            s."Name" AS student_name,
            COALESCE(s."Last_week_Submissions__c", 0) AS last_week_subs,
            COALESCE(s."Last_week_Interviews__c", 0) AS last_week_ints,
            CASE WHEN s."Verbal_Confirmation_Date__c" >= DATE_TRUNC('week', CURRENT_DATE) - INTERVAL '1 week'
                  AND s."Verbal_Confirmation_Date__c" < DATE_TRUNC('week', CURRENT_DATE)
                 THEN 1 ELSE 0 END AS confirmed
        FROM "Student__c" s
        LEFT JOIN "Manager__c" m ON s."Manager__c" = m."Id"
        WHERE s."Student_Marketing_Status__c" = 'In Market'
        ORDER BY s."Offshore_Manager_Name__c", m."Name", s."Recruiter_Name__c", s."Name"
    """)

    async with async_session() as session:
        result = await session.execute(sql)
        rows = result.fetchall()

    mgr_data = {}
    for offshore_mgr, bu_name, recruiter, student_name, subs, ints, confirmed in rows:
        mgr = offshore_mgr or "NAN"
        bu = bu_name or "Unknown"
        rec = recruiter or "NAN"

        if mgr not in mgr_data:
            mgr_data[mgr] = {"bus": {}, "total_subs": 0, "total_ints": 0, "total_conf": 0, "in_market": 0}
        mgr_data[mgr]["total_subs"] += int(subs)
        mgr_data[mgr]["total_ints"] += int(ints)
        mgr_data[mgr]["total_conf"] += int(confirmed)
        mgr_data[mgr]["in_market"] += 1

        if bu not in mgr_data[mgr]["bus"]:
            mgr_data[mgr]["bus"][bu] = {}
        if rec not in mgr_data[mgr]["bus"][bu]:
            mgr_data[mgr]["bus"][bu][rec] = []

        mgr_data[mgr]["bus"][bu][rec].append({
            "name": student_name or "Unknown",
            "subs": int(subs),
            "ints": int(ints),
            "conf": int(confirmed),
        })

    output_rows = []
    for mgr, data in sorted(mgr_data.items()):
        im = data["in_market"]
        target = round((data["total_subs"] / (im * 2 * 5) * 100), 2) if im > 0 else 0
        lines = []
        lines.append(f"\U0001f4ca Last Week Report\n")
        lines.append(f"*{mgr}*")
        lines.append(f"\U0001f4e4 Submissions Count: {data['total_subs']}")
        lines.append(f"\U0001f3a4 Interview Count: {data['total_ints']}")
        lines.append(f"\U0001f465 In Market Count: {im}")
        lines.append(f"\U0001f3af %% Target: {target}")
        lines.append(f"✅ Conformations: {data['total_conf']}")

        for bu, recruiters in sorted(data["bus"].items()):
            lines.append(f"\n\U0001f3e2 *BU: {bu}*")
            for rec, students in sorted(recruiters.items()):
                lines.append(f"\U0001f465 Recruiter: *{rec}*")
                for st in students:
                    lines.append(
                        f"\U0001f464 Student: *{st['name']}* | \U0001f4e4 Submissions: {st['subs']} "
                        f"| \U0001f3a4 Interviews: {st['ints']} | ✅ *Conformations:* {st['conf']}"
                    )

        output_rows.append({"Offshore Manager Name": mgr, "Message": "\n".join(lines)})

    return _xlsx_bytes(output_rows, ["Offshore Manager Name", "Message"], "Recruiter Performance OM")


# ────────────────────────────────────────────────────────────
# Registry — maps report_type to handler
# ────────────────────────────────────────────────────────────

REPORT_REGISTRY = {
    "premarketing_bu": {
        "label": "PreMarketing BU-wise",
        "category": "daily",
        "handler": premarketing_bu,
    },
    "yesterday_submissions_bu": {
        "label": "Yesterday Submissions BU-wise",
        "category": "daily",
        "handler": yesterday_submissions_bu,
    },
    "yesterday_submissions_offshore": {
        "label": "Yesterday Submissions Offshore Manager-wise",
        "category": "daily",
        "handler": yesterday_submissions_offshore,
    },
    "no_submissions_3days_bu": {
        "label": "Last 3 Days No Submissions BU-wise",
        "category": "daily",
        "handler": no_submissions_3days_bu,
    },
    "no_submissions_3days_offshore": {
        "label": "Last 3 Days No Submissions Offshore Manager-wise",
        "category": "daily",
        "handler": no_submissions_3days_offshore,
    },
    "interview_mandatory_fields_bu": {
        "label": "Interview Mandatory Fields BU-wise",
        "category": "daily",
        "handler": interview_mandatory_fields_bu,
    },
    "no_interviews_2weeks_bu": {
        "label": "Last 2 Weeks No Interviews BU-wise",
        "category": "weekly",
        "handler": no_interviews_2weeks_bu,
    },
    "no_interviews_2weeks_offshore": {
        "label": "Last 2 Weeks No Interviews Offshore Manager-wise",
        "category": "weekly",
        "handler": no_interviews_2weeks_offshore,
    },
    "last_week_performance_bu": {
        "label": "Last Week Submissions & Interviews BU-wise",
        "category": "weekly",
        "handler": last_week_performance_bu,
    },
    "last_week_performance_offshore": {
        "label": "Last Week Submissions & Interviews Offshore Manager-wise",
        "category": "weekly",
        "handler": last_week_performance_offshore,
    },
    "recruiter_performance_bu": {
        "label": "Recruiter Last Week Performance BU-wise",
        "category": "weekly",
        "handler": recruiter_performance_bu,
    },
    "recruiter_performance_offshore": {
        "label": "Recruiter Last Week Performance Offshore Manager-wise",
        "category": "weekly",
        "handler": recruiter_performance_offshore,
    },
}
