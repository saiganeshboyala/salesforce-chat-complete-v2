"""
File uploads — per-user CSV/XLSX parsing for ad-hoc analysis in chat.

Storage:
  {data_dir}/users/{username}/uploads/{upload_id}/
    source.csv | source.xlsx        (original bytes)
    meta.json                       (metadata + row summary + preview)

The chat layer can attach an upload_id to a question; when it does,
`load_context_text()` returns a compact textual summary (headers + first
rows + row count) that the AI engine prepends to the user prompt.
"""
import csv, io, json, logging, re, uuid
from datetime import datetime
from pathlib import Path
from app.config import settings

logger = logging.getLogger(__name__)

MAX_PREVIEW_ROWS = 30
MAX_PROMPT_ROWS = 50
MAX_PROMPT_CHARS = 6000
SUPPORTED_EXT = (".csv", ".xlsx", ".xls")


# ── Paths ──────────────────────────────────────────────

def _user_dir(username: str) -> Path:
    d = Path(settings.data_dir) / "users" / username / "uploads"
    d.mkdir(parents=True, exist_ok=True)
    return d


def _upload_dir(username: str, upload_id: str) -> Path:
    safe = re.sub(r"[^A-Za-z0-9_\-]", "_", upload_id)
    return _user_dir(username) / safe


def _meta_path(username: str, upload_id: str) -> Path:
    return _upload_dir(username, upload_id) / "meta.json"


# ── Parsers ────────────────────────────────────────────

def _parse_csv(raw: bytes) -> tuple[list[str], list[dict], int]:
    text = raw.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    if not rows:
        return [], [], 0
    headers = [h.strip() or f"col_{i}" for i, h in enumerate(rows[0])]
    records = []
    for r in rows[1:]:
        records.append({h: (r[i] if i < len(r) else "") for i, h in enumerate(headers)})
    return headers, records, len(records)


def _parse_xlsx(raw: bytes) -> tuple[list[str], list[dict], int]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return [], [], 0
    headers = [
        (str(h).strip() if h is not None else f"col_{i}") or f"col_{i}"
        for i, h in enumerate(header_row)
    ]
    records = []
    count = 0
    for row in rows_iter:
        count += 1
        records.append({h: ("" if row[i] is None else row[i]) for i, h in enumerate(headers) if i < len(row)})
    wb.close()
    return headers, records, count


def parse_upload(filename: str, raw: bytes) -> tuple[list[str], list[dict], int]:
    ext = Path(filename).suffix.lower()
    if ext == ".csv":
        return _parse_csv(raw)
    if ext in (".xlsx", ".xls"):
        return _parse_xlsx(raw)
    raise ValueError(f"Unsupported file type: {ext}")


# ── CRUD ───────────────────────────────────────────────

def save_upload(username: str, filename: str, raw: bytes) -> dict:
    ext = Path(filename).suffix.lower()
    if ext not in SUPPORTED_EXT:
        raise ValueError(f"Unsupported file type: {ext}")

    headers, records, total = parse_upload(filename, raw)
    if not headers:
        raise ValueError("File appears to be empty")

    upload_id = f"up_{uuid.uuid4().hex[:10]}"
    folder = _upload_dir(username, upload_id)
    folder.mkdir(parents=True, exist_ok=True)

    # Store original file for re-download
    src_name = "source" + ext
    (folder / src_name).write_bytes(raw)

    preview = records[:MAX_PREVIEW_ROWS]
    # JSON-safe preview (stringify anything weird)
    safe_preview = [{k: ("" if v is None else v) for k, v in r.items()} for r in preview]

    meta = {
        "id": upload_id,
        "filename": filename,
        "size_bytes": len(raw),
        "uploaded_at": datetime.now().isoformat(),
        "source_file": src_name,
        "headers": headers,
        "row_count": total,
        "preview": safe_preview,
    }

    with open(_meta_path(username, upload_id), "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2, default=str)

    return meta


def list_uploads(username: str) -> list[dict]:
    try:
        d = _user_dir(username)
    except Exception as e:
        logger.warning(f"list_uploads: cannot access user dir for {username}: {e}")
        return []
    out = []
    try:
        entries = list(d.iterdir())
    except Exception as e:
        logger.warning(f"list_uploads: iterdir failed for {d}: {e}")
        return []
    for folder in entries:
        try:
            if not folder.is_dir():
                continue
            meta_path = folder / "meta.json"
            if not meta_path.exists():
                continue
            with open(meta_path, encoding="utf-8") as f:
                m = json.load(f)
            headers = m.get("headers") or []
            out.append({
                "id": m.get("id", folder.name),
                "filename": m.get("filename") or folder.name,
                "row_count": int(m.get("row_count") or 0),
                "size_bytes": int(m.get("size_bytes") or 0),
                "uploaded_at": str(m.get("uploaded_at") or ""),
                "headers": list(headers),
            })
        except Exception as e:
            logger.warning(f"Failed to read upload {folder}: {e}")
    out.sort(key=lambda x: x.get("uploaded_at", ""), reverse=True)
    return out


def get_upload(username: str, upload_id: str) -> dict | None:
    p = _meta_path(username, upload_id)
    if not p.exists():
        return None
    with open(p, encoding="utf-8") as f:
        return json.load(f)


def delete_upload(username: str, upload_id: str) -> bool:
    folder = _upload_dir(username, upload_id)
    if not folder.exists():
        return False
    for f in folder.iterdir():
        try:
            f.unlink()
        except Exception:
            pass
    try:
        folder.rmdir()
    except Exception:
        pass
    return True


def load_context_text(username: str, upload_id: str) -> str | None:
    """
    Return a compact textual summary of an uploaded file, suitable
    for injection into the AI prompt. Truncates aggressively.
    """
    meta = get_upload(username, upload_id)
    if not meta:
        return None
    headers = meta.get("headers", [])
    preview = meta.get("preview", [])[:MAX_PROMPT_ROWS]
    lines = [
        f'Attached file: "{meta.get("filename")}" ({meta.get("row_count", 0)} rows)',
        "Columns: " + ", ".join(headers),
        "Data preview:",
        " | ".join(headers),
    ]
    for r in preview:
        lines.append(" | ".join(str(r.get(h, "")) for h in headers))
    text = "\n".join(lines)
    if len(text) > MAX_PROMPT_CHARS:
        text = text[:MAX_PROMPT_CHARS] + "\n…(truncated)"
    return text
