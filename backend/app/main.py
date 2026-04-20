"""
Salesforce Data Chat — API Server with Authentication
"""
import csv, io, logging, re, threading
from datetime import datetime, timedelta
from contextlib import asynccontextmanager
from pathlib import Path

from fastapi import FastAPI, HTTPException, Query, Depends, UploadFile, File, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, RedirectResponse
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel
from slowapi import Limiter
from slowapi.util import get_remote_address
from slowapi.errors import RateLimitExceeded
from starlette.responses import JSONResponse

from app.config import settings
from app.models.schemas import ChatRequest, ChatResponse
from app.chat.engine import ChatEngine
from app.chat.memory import update_feedback, get_stats as memory_stats, get_user_history, get_user_stats
from app.chat.sessions import (
    list_sessions as list_user_sessions,
    load_session as load_user_session,
    delete_session as delete_user_session,
    search_sessions as search_user_sessions,
    toggle_pin as toggle_session_pin,
)
from app import schedules as sched
from app import uploads as up
from app import connectors as conn
from app import audit
from app import dashboard_config as dash_cfg
from app import compare as compare_mod
from app import alerts as alerts_mod
from app import annotations as annotations_mod
from app import reports as reports_mod
from app import analytics as analytics_mod
from app.connectors import gmail as gmail_conn
from app.connectors import google_oauth as google_oauth
from app.salesforce.schema import get_schema, discover_schema, get_relationships
from app.salesforce.soql_executor import execute_soql
from app.auth_users import (
    authenticate_user, create_token, get_current_user, get_optional_user,
    create_user, list_users, delete_user, change_password, decode_token,
)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)
chat_engine = None


# ── Lifecycle ──────────────────────────────────────────

async def refresh_schema_task():
    from app.salesforce.auth import ensure_authenticated
    creds = await ensure_authenticated()
    await discover_schema(creds.instance_url, creds.access_token)


def start_scheduler():
    sync_time = settings.auto_sync_time.strip()
    if not sync_time:
        return
    try:
        hour, minute = map(int, sync_time.split(":"))
    except ValueError:
        return
    logger.info(f"Schema refresh scheduled daily at {sync_time}")
    def loop():
        import time, asyncio
        while True:
            now = datetime.now()
            target = now.replace(hour=hour, minute=minute, second=0)
            if target <= now:
                target += timedelta(days=1)
            time.sleep((target - now).total_seconds())
            asyncio.run(refresh_schema_task())
    threading.Thread(target=loop, daemon=True).start()


@asynccontextmanager
async def lifespan(app):
    global chat_engine
    schema = get_schema()
    if not schema:
        try:
            await refresh_schema_task()
        except Exception as e:
            logger.error(f"Schema discovery failed: {e}")
    chat_engine = ChatEngine()
    logger.info(f"Ready — {len(get_schema())} objects | Auth + SOQL + RAG + Learning")
    start_scheduler()
    sched.start_runner()
    yield
    sched.stop_runner()


limiter = Limiter(key_func=get_remote_address)

app = FastAPI(title="Salesforce Data Chat", lifespan=lifespan)
app.state.limiter = limiter

@app.exception_handler(RateLimitExceeded)
async def _rate_limit_handler(request: Request, exc: RateLimitExceeded):
    return JSONResponse(status_code=429, content={"detail": "Too many requests — please slow down"})

app.add_middleware(CORSMiddleware, allow_origins=settings.cors_origins, allow_credentials=True, allow_methods=["*"], allow_headers=["*"])


_SOQL_WRITE_RE = re.compile(r"\b(INSERT|UPDATE|DELETE|UPSERT|MERGE|UNDELETE)\b", re.IGNORECASE)

def _assert_read_only_soql(soql: str):
    if _SOQL_WRITE_RE.search(soql):
        raise HTTPException(400, "Only SELECT queries are allowed")


# ── Auth Endpoints ─────────────────────────────────────

class LoginRequest(BaseModel):
    username: str
    password: str

class RegisterRequest(BaseModel):
    username: str
    password: str
    name: str
    role: str | None = "user"

class ChangePasswordRequest(BaseModel):
    old_password: str
    new_password: str

@app.post("/api/auth/login")
@limiter.limit("10/minute")
async def login(req: LoginRequest, request: Request):
    user = authenticate_user(req.username, req.password)
    ip = request.client.host if request.client else ""
    if not user:
        audit.log_action(req.username, "login_failed", {"reason": "bad_credentials"}, ip)
        raise HTTPException(401, "Invalid username or password")
    token = create_token(user["username"], user.get("role", "user"))
    audit.log_action(user["username"], "login", {}, ip)
    return {
        "token": token,
        "user": {"username": user["username"], "name": user.get("name", ""), "role": user.get("role", "user")},
    }

@app.post("/api/auth/register")
async def register(req: RegisterRequest, current_user=Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(403, "Only admins can create users")
    role = req.role if req.role in ("admin", "user") else "user"
    try:
        user = create_user(req.username, req.password, req.name, role=role)
    except ValueError as e:
        raise HTTPException(400, str(e))
    if not user:
        raise HTTPException(400, "Username already exists")
    return {"status": "created", "username": req.username}

@app.get("/api/auth/me")
async def me(current_user=Depends(get_current_user)):
    stats = get_user_stats(current_user["username"])
    return {**current_user, **stats}

@app.post("/api/auth/change-password")
async def change_pwd(req: ChangePasswordRequest, current_user=Depends(get_current_user)):
    user = authenticate_user(current_user["username"], req.old_password)
    if not user:
        raise HTTPException(400, "Current password is wrong")
    try:
        change_password(current_user["username"], req.new_password)
    except ValueError as e:
        raise HTTPException(400, str(e))
    return {"status": "changed"}

@app.get("/api/auth/users")
async def get_users(current_user=Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(403, "Admin only")
    return list_users()

@app.delete("/api/auth/users/{username}")
async def remove_user(username: str, current_user=Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(403, "Admin only")
    if not delete_user(username):
        raise HTTPException(400, "Cannot delete this user")
    return {"status": "deleted"}


# ── Chat (with optional auth) ─────────────────────────

@app.post("/api/chat", response_model=ChatResponse)
@limiter.limit("30/minute")
async def chat(req: ChatRequest, request: Request, user=Depends(get_optional_user)):
    if not chat_engine:
        raise HTTPException(503, "Not initialized")
    try:
        username = user["username"] if user else None
        question = req.question
        if req.attachment_id and username:
            ctx = up.load_context_text(username, req.attachment_id)
            if ctx:
                question = f"{ctx}\n\n---\n\nUser question: {req.question}"
        result = await chat_engine.answer(req.session_id, question, username=username)
        audit.log_action(
            username, "chat_question",
            {"question": (req.question or "")[:100]},
            request.client.host if request.client else "",
        )
        return ChatResponse(answer=result["answer"], soql=result.get("soql"), data=result.get("data"))
    except Exception as e:
        logger.error(f"Chat error: {e}")
        raise HTTPException(500, str(e))


@app.post("/api/chat/stream")
@limiter.limit("30/minute")
async def chat_stream(req: ChatRequest, request: Request, user=Depends(get_optional_user)):
    """Server-Sent Events stream for chat answers (word-by-word)."""
    if not chat_engine:
        raise HTTPException(503, "Not initialized")

    import json as _json

    username = user["username"] if user else None
    question = req.question
    if req.attachment_id and username:
        ctx = up.load_context_text(username, req.attachment_id)
        if ctx:
            question = f"{ctx}\n\n---\n\nUser question: {req.question}"

    audit.log_action(
        username, "chat_question",
        {"question": (req.question or "")[:100]},
        request.client.host if request.client else "",
    )

    async def event_gen():
        try:
            async for event in chat_engine.answer_stream(req.session_id, question, username=username):
                yield f"data: {_json.dumps(event, default=str)}\n\n"
        except Exception as e:
            logger.error(f"Chat stream error: {e}")
            err = {"type": "error", "data": str(e)}
            yield f"data: {_json.dumps(err)}\n\n"

    return StreamingResponse(
        event_gen(),
        media_type="text/event-stream",
        headers={
            "Cache-Control": "no-cache",
            "X-Accel-Buffering": "no",
            "Connection": "keep-alive",
        },
    )


@app.get("/api/welcome")
async def welcome():
    if not chat_engine:
        return {"answer": "Initializing...", "data": None}
    return chat_engine.get_welcome()


# ── Feedback ───────────────────────────────────────────

class FeedbackRequest(BaseModel):
    question: str
    feedback: str

@app.post("/api/feedback")
async def feedback(req: FeedbackRequest, request: Request, user=Depends(get_optional_user)):
    update_feedback(req.question, req.feedback)
    audit.log_action(
        (user or {}).get("username"),
        "feedback",
        {"question": (req.question or "")[:50], "type": req.feedback},
        request.client.host if request.client else "",
    )
    return {"status": "saved", "feedback": req.feedback}

@app.get("/api/learning-stats")
async def learning_stats():
    return memory_stats()


# ── User History ───────────────────────────────────────

@app.get("/api/history")
async def history(current_user=Depends(get_current_user)):
    return get_user_history(current_user["username"])


# ── Chat Sessions ──────────────────────────────────────

@app.get("/api/sessions")
async def sessions_list(q: str | None = None, current_user=Depends(get_current_user)):
    username = current_user["username"]
    if q:
        return search_user_sessions(username, q)
    return list_user_sessions(username)


@app.get("/api/sessions/{session_id}")
async def sessions_get(session_id: str, current_user=Depends(get_current_user)):
    return load_user_session(current_user["username"], session_id)


@app.delete("/api/sessions/{session_id}")
async def sessions_delete(session_id: str, current_user=Depends(get_current_user)):
    ok = delete_user_session(current_user["username"], session_id)
    if not ok:
        raise HTTPException(404, "Session not found")
    return {"status": "deleted"}


@app.post("/api/sessions/{session_id}/pin")
async def sessions_pin(session_id: str, current_user=Depends(get_current_user)):
    pinned = toggle_session_pin(current_user["username"], session_id)
    return {"pinned": pinned}


# ── Audit Log ──────────────────────────────────────────

@app.get("/api/audit")
async def audit_list(
    user: str | None = None,
    action: str | None = None,
    start: str | None = None,
    end: str | None = None,
    page: int = 1,
    page_size: int = 50,
    current_user=Depends(get_current_user),
):
    if current_user["role"] != "admin":
        raise HTTPException(403, "Admin only")
    return audit.query_log(user=user, action=action, start=start, end=end, page=page, page_size=page_size)


# ── Admin: password reset ──────────────────────────────

class AdminResetPasswordRequest(BaseModel):
    username: str
    new_password: str

@app.post("/api/auth/admin-reset-password")
async def admin_reset_password(req: AdminResetPasswordRequest, current_user=Depends(get_current_user)):
    if current_user["role"] != "admin":
        raise HTTPException(403, "Admin only")
    try:
        ok = change_password(req.username, req.new_password)
    except ValueError as e:
        raise HTTPException(400, str(e))
    if not ok:
        raise HTTPException(404, "User not found")
    return {"status": "reset"}


# ── Dashboard ─────────────────────────────────────────

class DashboardConfigRequest(BaseModel):
    widgets: list


@app.get("/api/dashboard/config")
async def dashboard_config_get(current_user=Depends(get_current_user)):
    return dash_cfg.load_config(current_user["username"])


@app.post("/api/dashboard/config")
async def dashboard_config_save(req: DashboardConfigRequest, current_user=Depends(get_current_user)):
    try:
        return dash_cfg.save_config(current_user["username"], {"widgets": req.widgets})
    except ValueError as e:
        raise HTTPException(400, str(e))


class CompareRequest(BaseModel):
    query1: str | None = None
    query2: str | None = None
    label1: str | None = None
    label2: str | None = None
    question: str | None = None


@app.post("/api/compare")
async def compare_endpoint(req: CompareRequest, user=Depends(get_optional_user)):
    try:
        if req.question:
            return await compare_mod.run_compare_question(req.question)
        if not req.query1 or not req.query2:
            raise HTTPException(400, "Either question or both query1 and query2 required")
        return await compare_mod.run_compare(
            req.query1, req.query2,
            req.label1 or "Period 1", req.label2 or "Period 2",
        )
    except ValueError as e:
        raise HTTPException(400, str(e))
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Compare error: {e}")
        raise HTTPException(500, str(e))


# ── Alert Rules ──────────────────────────────────────

class AlertRuleRequest(BaseModel):
    name: str
    soql: str
    condition: str = "greater_than"
    threshold: float = 0
    frequency: str = "daily"
    enabled: bool = True


class AlertRulePatch(BaseModel):
    name: str | None = None
    soql: str | None = None
    condition: str | None = None
    threshold: float | None = None
    frequency: str | None = None
    enabled: bool | None = None


@app.get("/api/alerts")
async def alerts_list(current_user=Depends(get_current_user)):
    return {"rules": alerts_mod.list_rules(current_user["username"])}


@app.post("/api/alerts")
async def alerts_create(req: AlertRuleRequest, current_user=Depends(get_current_user)):
    _assert_read_only_soql(req.soql)
    try:
        rule = alerts_mod.create_rule(current_user["username"], req.model_dump())
        return rule
    except ValueError as e:
        raise HTTPException(400, str(e))


@app.patch("/api/alerts/{rule_id}")
async def alerts_update(rule_id: str, req: AlertRulePatch, current_user=Depends(get_current_user)):
    patch = {k: v for k, v in req.model_dump().items() if v is not None}
    try:
        return alerts_mod.update_rule(current_user["username"], rule_id, patch)
    except KeyError:
        raise HTTPException(404, "rule not found")
    except ValueError as e:
        raise HTTPException(400, str(e))


@app.delete("/api/alerts/{rule_id}")
async def alerts_delete(rule_id: str, current_user=Depends(get_current_user)):
    alerts_mod.delete_rule(current_user["username"], rule_id)
    return {"status": "deleted"}


@app.post("/api/alerts/{rule_id}/check")
async def alerts_check_one(rule_id: str, current_user=Depends(get_current_user)):
    try:
        return await alerts_mod.check_rule(current_user["username"], rule_id)
    except KeyError:
        raise HTTPException(404, "rule not found")
    except Exception as e:
        raise HTTPException(500, str(e))


@app.post("/api/alerts/check")
async def alerts_check_all(current_user=Depends(get_current_user)):
    return {"checked": await alerts_mod.check_all(current_user["username"])}


@app.get("/api/alerts/history")
async def alerts_history(current_user=Depends(get_current_user)):
    return {"entries": alerts_mod.list_history(current_user["username"])}


# ── Data Annotations ─────────────────────────────────

class AnnotationRequest(BaseModel):
    record_id: str
    record_name: str | None = None
    object_type: str | None = None
    text: str
    tags: list[str] | None = None


class AnnotationPatch(BaseModel):
    text: str | None = None
    tags: list[str] | None = None


class AnnotationLookupRequest(BaseModel):
    record_ids: list[str]


@app.get("/api/annotations")
async def annotations_list(
    record_id: str | None = None,
    tag: str | None = None,
    q: str | None = None,
    current_user=Depends(get_current_user),
):
    notes = annotations_mod.list_notes(current_user["username"], record_id=record_id, tag=tag, q=q)
    return {"notes": notes}


@app.post("/api/annotations")
async def annotations_create(req: AnnotationRequest, current_user=Depends(get_current_user)):
    try:
        return annotations_mod.create_note(current_user["username"], req.model_dump())
    except ValueError as e:
        raise HTTPException(400, str(e))


@app.patch("/api/annotations/{note_id}")
async def annotations_update(note_id: str, req: AnnotationPatch, current_user=Depends(get_current_user)):
    patch = {k: v for k, v in req.model_dump().items() if v is not None}
    try:
        return annotations_mod.update_note(current_user["username"], note_id, patch)
    except KeyError:
        raise HTTPException(404, "note not found")


@app.delete("/api/annotations/{note_id}")
async def annotations_delete(note_id: str, current_user=Depends(get_current_user)):
    try:
        annotations_mod.delete_note(current_user["username"], note_id)
        return {"status": "deleted"}
    except KeyError:
        raise HTTPException(404, "note not found")


@app.get("/api/annotations/tags")
async def annotations_tags(current_user=Depends(get_current_user)):
    return {"tags": annotations_mod.list_tags(current_user["username"])}


@app.post("/api/annotations/lookup")
async def annotations_lookup(req: AnnotationLookupRequest, current_user=Depends(get_current_user)):
    return {"map": annotations_mod.get_for_records(current_user["username"], req.record_ids)}


# ── Report Builder ────────────────────────────────────

class ReportFilter(BaseModel):
    field: str
    operator: str
    value: str | int | float | bool | None = None


class ReportRequest(BaseModel):
    name: str
    description: str | None = None
    object: str
    fields: list[str] = []
    filters: list[ReportFilter] = []
    groupBy: str | None = None
    chartType: str = "none"
    sortBy: str | None = None
    sortDir: str = "asc"
    limit: int | None = 200


class ReportPatch(BaseModel):
    name: str | None = None
    description: str | None = None
    object: str | None = None
    fields: list[str] | None = None
    filters: list[ReportFilter] | None = None
    groupBy: str | None = None
    chartType: str | None = None
    sortBy: str | None = None
    sortDir: str | None = None
    limit: int | None = None


class ReportRunRequest(BaseModel):
    config: ReportRequest | None = None


class ReportSuggestRequest(BaseModel):
    prompt: str


@app.get("/api/reports")
async def reports_list(current_user=Depends(get_current_user)):
    return {"reports": reports_mod.list_reports(current_user["username"])}


@app.post("/api/reports")
async def reports_create(req: ReportRequest, current_user=Depends(get_current_user)):
    try:
        return reports_mod.create_report(current_user["username"], req.model_dump())
    except ValueError as e:
        raise HTTPException(400, str(e))


@app.get("/api/reports/{report_id}")
async def reports_get(report_id: str, current_user=Depends(get_current_user)):
    try:
        return reports_mod.get_report(current_user["username"], report_id)
    except KeyError:
        raise HTTPException(404, "report not found")


@app.patch("/api/reports/{report_id}")
async def reports_update(report_id: str, req: ReportPatch, current_user=Depends(get_current_user)):
    patch = {k: v for k, v in req.model_dump().items() if v is not None}
    try:
        return reports_mod.update_report(current_user["username"], report_id, patch)
    except KeyError:
        raise HTTPException(404, "report not found")
    except ValueError as e:
        raise HTTPException(400, str(e))


@app.delete("/api/reports/{report_id}")
async def reports_delete(report_id: str, current_user=Depends(get_current_user)):
    try:
        reports_mod.delete_report(current_user["username"], report_id)
        return {"status": "deleted"}
    except KeyError:
        raise HTTPException(404, "report not found")


@app.post("/api/reports/{report_id}/run")
async def reports_run(report_id: str, current_user=Depends(get_current_user)):
    try:
        return await reports_mod.run_report(current_user["username"], report_id=report_id)
    except KeyError:
        raise HTTPException(404, "report not found")
    except ValueError as e:
        raise HTTPException(400, str(e))


@app.post("/api/reports/preview")
async def reports_preview(req: ReportRequest, current_user=Depends(get_current_user)):
    try:
        return await reports_mod.run_report(current_user["username"], config=req.model_dump())
    except ValueError as e:
        raise HTTPException(400, str(e))


@app.post("/api/reports/suggest")
async def reports_suggest(req: ReportSuggestRequest, current_user=Depends(get_current_user)):
    if not req.prompt or not req.prompt.strip():
        raise HTTPException(400, "prompt is required")
    try:
        return await reports_mod.suggest_report(req.prompt.strip())
    except ValueError as e:
        raise HTTPException(400, str(e))


@app.get("/api/dashboard/widget")
async def dashboard_widget(q: str = Query(...), current_user=Depends(get_current_user)):
    """Execute a single widget's SOQL and return its data."""
    _assert_read_only_soql(q)
    try:
        result = await execute_soql(q)
        if "error" in result:
            raise HTTPException(400, result["error"])
        for r in result.get("records", []):
            r.pop("attributes", None)
        return result
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


@app.get("/api/dashboard")
async def dashboard():
    queries = {
        "total_students": "SELECT COUNT() FROM Student__c",
        "students_in_market": "SELECT COUNT() FROM Student__c WHERE Student_Marketing_Status__c = 'In Market'",
        "verbal_confirmations": "SELECT COUNT() FROM Student__c WHERE Student_Marketing_Status__c = 'Verbal Confirmation'",
        "project_started": "SELECT COUNT() FROM Student__c WHERE Student_Marketing_Status__c = 'Project Started'",
        "exits": "SELECT COUNT() FROM Student__c WHERE Student_Marketing_Status__c = 'Exit'",
        "total_accounts": "SELECT COUNT() FROM Account",
        "total_contacts": "SELECT COUNT() FROM Contact",
        "status_breakdown": "SELECT Student_Marketing_Status__c, COUNT(Id) cnt FROM Student__c GROUP BY Student_Marketing_Status__c ORDER BY COUNT(Id) DESC",
    }
    metrics = {}
    for key, soql in queries.items():
        try:
            result = await execute_soql(soql)
            if "error" not in result:
                if key == "status_breakdown":
                    records = result.get("records", [])
                    for r in records:
                        r.pop("attributes", None)
                    metrics[key] = records
                else:
                    metrics[key] = result.get("totalSize", 0)
        except Exception:
            metrics[key] = 0
    return metrics


# ── Export ─────────────────────────────────────────────

def _records_to_xlsx_bytes(records: list[dict], headers: list[str]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Salesforce Export"

    header_font = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    header_fill = PatternFill(start_color="E8734A", end_color="E8734A", fill_type="solid")
    header_align = Alignment(horizontal="left", vertical="center")

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    for row_idx, rec in enumerate(records, start=2):
        for col_idx, h in enumerate(headers, start=1):
            val = rec.get(h)
            if isinstance(val, (dict, list)):
                val = str(val)
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-size columns based on sampled content
    sample = records[:200]
    for col_idx, h in enumerate(headers, start=1):
        max_len = len(str(h))
        for rec in sample:
            v = rec.get(h)
            if v is None:
                continue
            max_len = max(max_len, min(len(str(v)), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 3

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


class PdfExportRequest(BaseModel):
    title: str | None = None
    question: str = ""
    answer: str = ""
    soql: str | None = None
    records: list | None = None
    chart_type: str | None = None


@app.post("/api/export/pdf")
async def export_pdf(req: PdfExportRequest, request: Request, user=Depends(get_optional_user)):
    from app.pdf_export import build_pdf
    try:
        username = (user or {}).get("username", "user")
        audit.log_action(
            username, "export",
            {"format": "pdf", "query": (req.soql or "")[:100]},
            request.client.host if request.client else "",
        )
        pdf_bytes = build_pdf(
            title=req.title or "Salesforce Data Report",
            question=req.question or "",
            answer=req.answer or "",
            soql=req.soql,
            records=req.records or [],
            username=username,
        )
        safe_name = (req.title or "salesforce_report").lower()
        safe_name = "".join(c if c.isalnum() else "_" for c in safe_name)[:60] or "report"
        return StreamingResponse(
            iter([pdf_bytes]),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename={safe_name}.pdf"},
        )
    except Exception as e:
        logger.error(f"PDF export error: {e}")
        raise HTTPException(500, f"PDF generation failed: {e}")


@app.get("/api/export")
async def export_data(q: str = Query(...), format: str = Query("csv"), current_user=Depends(get_current_user)):
    _assert_read_only_soql(q)
    fmt = (format or "csv").lower()
    if fmt not in ("csv", "xlsx"):
        raise HTTPException(400, "format must be csv or xlsx")
    try:
        result = await execute_soql(q)
        if "error" in result:
            raise HTTPException(400, result["error"])
        records = result.get("records", [])
        for r in records:
            r.pop("attributes", None)
        if not records:
            raise HTTPException(404, "No records")
        headers = list(records[0].keys())

        if fmt == "xlsx":
            data = _records_to_xlsx_bytes(records, headers)
            return StreamingResponse(
                iter([data]),
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": "attachment; filename=salesforce_export.xlsx"},
            )

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=headers, extrasaction="ignore")
        writer.writeheader()
        for r in records:
            writer.writerow(r)
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=salesforce_export.csv"},
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, str(e))


# ── Scheduled Reports ──────────────────────────────────

class ScheduleCreateRequest(BaseModel):
    name: str | None = None
    question: str | None = None
    soql: str
    frequency: str = "daily"  # daily | weekly | monthly
    time: str = "09:00"       # HH:MM
    weekday: int | None = None
    day_of_month: int | None = None
    recipients: list[str] = []

class ScheduleUpdateRequest(BaseModel):
    name: str | None = None
    question: str | None = None
    soql: str | None = None
    frequency: str | None = None
    time: str | None = None
    weekday: int | None = None
    day_of_month: int | None = None
    recipients: list[str] | None = None
    enabled: bool | None = None


@app.get("/api/schedules")
async def schedules_list(current_user=Depends(get_current_user)):
    return sched.list_schedules(current_user["username"])


@app.post("/api/schedules")
async def schedules_create(req: ScheduleCreateRequest, current_user=Depends(get_current_user)):
    if not req.soql or not req.soql.strip():
        raise HTTPException(400, "soql is required")
    _assert_read_only_soql(req.soql)
    if req.frequency not in ("daily", "weekly", "monthly"):
        raise HTTPException(400, "frequency must be daily, weekly, or monthly")
    return sched.create_schedule(current_user["username"], req.model_dump(exclude_none=True))


@app.patch("/api/schedules/{schedule_id}")
async def schedules_update(schedule_id: str, req: ScheduleUpdateRequest, current_user=Depends(get_current_user)):
    updated = sched.update_schedule(current_user["username"], schedule_id, req.model_dump(exclude_none=True))
    if not updated:
        raise HTTPException(404, "Schedule not found")
    return updated


@app.delete("/api/schedules/{schedule_id}")
async def schedules_delete(schedule_id: str, current_user=Depends(get_current_user)):
    if not sched.delete_schedule(current_user["username"], schedule_id):
        raise HTTPException(404, "Schedule not found")
    return {"status": "deleted"}


@app.post("/api/schedules/{schedule_id}/run")
async def schedules_run_now(schedule_id: str, current_user=Depends(get_current_user)):
    meta = await sched.run_schedule_now(current_user["username"], schedule_id)
    if not meta:
        raise HTTPException(404, "Schedule not found")
    return meta


@app.get("/api/schedules/{schedule_id}/runs")
async def schedules_runs(schedule_id: str, current_user=Depends(get_current_user)):
    return sched.list_runs(current_user["username"], schedule_id)


# ── Uploads ────────────────────────────────────────────

MAX_UPLOAD_BYTES = 15 * 1024 * 1024  # 15 MB

@app.post("/api/uploads")
async def uploads_create(file: UploadFile = File(...), current_user=Depends(get_current_user)):
    raw = await file.read()
    if len(raw) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, f"File too large (max {MAX_UPLOAD_BYTES // (1024*1024)} MB)")
    try:
        meta = up.save_upload(current_user["username"], file.filename or "upload", raw)
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        logger.error(f"Upload failed: {e}")
        raise HTTPException(500, f"Failed to parse file: {e}")
    return meta


@app.get("/api/uploads")
async def uploads_list(current_user=Depends(get_current_user)):
    return up.list_uploads(current_user["username"])


@app.get("/api/uploads/{upload_id}")
async def uploads_get(upload_id: str, current_user=Depends(get_current_user)):
    meta = up.get_upload(current_user["username"], upload_id)
    if not meta:
        raise HTTPException(404, "Upload not found")
    return meta


@app.delete("/api/uploads/{upload_id}")
async def uploads_delete(upload_id: str, current_user=Depends(get_current_user)):
    if not up.delete_upload(current_user["username"], upload_id):
        raise HTTPException(404, "Upload not found")
    return {"status": "deleted"}


# ── Schema ─────────────────────────────────────────────

@app.get("/api/overview")
async def overview():
    schema = get_schema()
    return {
        "total_objects": len(schema),
        "total_records": sum(s.get("record_count", 0) or 0 for s in schema.values()),
        "objects": {n: {"label": s.get("label"), "records": s.get("record_count"), "fields": len(s.get("fields", []))} for n, s in schema.items()},
    }

@app.get("/api/schema/relationships")
async def schema_relationships():
    return get_relationships()


@app.get("/api/schema/objects")
async def schema_objects():
    """Full schema: { object: { label, record_count, fields: [{name,type,label,filterable,groupable,sortable}] } }"""
    schema = get_schema()
    return {
        name: {
            "label": meta.get("label", name),
            "record_count": meta.get("record_count"),
            "fields": meta.get("fields", []),
        }
        for name, meta in schema.items()
    }


@app.get("/api/health")
async def health():
    schema = get_schema()
    providers = []
    if settings.anthropic_api_key: providers.append(f"Claude ({settings.claude_model})")
    if settings.grok_api_key: providers.append(f"Grok ({settings.grok_model})")
    if settings.openai_api_key: providers.append(f"OpenAI ({settings.openai_model})")
    return {"status": "healthy", "mode": "hybrid_soql_rag_learning_auth", "objects": len(schema), "ai_providers": providers, "learning": memory_stats()}

@app.post("/api/refresh-schema")
async def refresh_schema():
    try:
        await refresh_schema_task()
        return {"status": "refreshed", "objects": len(get_schema())}
    except Exception as e:
        raise HTTPException(500, str(e))


# ── Connectors ─────────────────────────────────────────

class GmailSendRequest(BaseModel):
    to: str
    subject: str = ""
    body: str = ""
    cc: str | None = None
    bcc: str | None = None


@app.get("/api/connectors")
async def connectors_list(current_user=Depends(get_current_user)):
    return conn.list_connectors(current_user["username"])


@app.post("/api/connectors/{connector_id}/disconnect")
async def connector_disconnect(connector_id: str, current_user=Depends(get_current_user)):
    ok = conn.disconnect(current_user["username"], connector_id)
    if not ok:
        raise HTTPException(404, "Connector not found or not connected")
    return {"status": "disconnected"}


def _user_from_query_token(token: str | None) -> dict:
    """
    For browser-redirect endpoints we can't carry an Authorization
    header, so the frontend passes ?token=<jwt>. This helper validates
    it the same way get_current_user does.
    """
    if not token:
        raise HTTPException(401, "Missing token")
    payload = decode_token(token)
    if not payload:
        raise HTTPException(401, "Invalid or expired token")
    username = payload.get("sub")
    if not username:
        raise HTTPException(401, "Invalid token")
    return {"username": username, "role": payload.get("role", "user")}


# Gmail OAuth entry / callback
@app.get("/api/connectors/gmail/auth")
async def gmail_auth(token: str | None = None):
    user = _user_from_query_token(token)
    if not google_oauth.is_configured():
        raise HTTPException(503, "Google OAuth not configured on server")
    try:
        url = gmail_conn.authorize_url(user["username"])
    except RuntimeError as e:
        raise HTTPException(503, str(e))
    return RedirectResponse(url)


@app.get("/api/connectors/google/callback")
async def google_callback(code: str | None = None, state: str | None = None,
                          error: str | None = None):
    """
    Shared Google OAuth callback — handles Gmail, Sheets, and Calendar.
    Redirects the browser back to the frontend with a status flag.
    """
    frontend = settings.frontend_origin.rstrip("/")

    if error:
        return RedirectResponse(f"{frontend}/?connector=error&reason={error}")
    if not code or not state:
        return RedirectResponse(f"{frontend}/?connector=error&reason=missing_params")

    consumed = google_oauth.consume_state(state)
    if not consumed:
        return RedirectResponse(f"{frontend}/?connector=error&reason=invalid_state")
    username, return_connector = consumed

    try:
        google_oauth.exchange_code(code, username)
    except Exception as e:
        logger.error(f"Google OAuth exchange failed: {e}")
        return RedirectResponse(f"{frontend}/?connector=error&reason=exchange_failed")

    return RedirectResponse(f"{frontend}/?connector=connected&id={return_connector}")


@app.post("/api/connectors/gmail/send")
async def gmail_send(req: GmailSendRequest, current_user=Depends(get_current_user)):
    try:
        result = gmail_conn.send_email(
            current_user["username"],
            to=req.to,
            subject=req.subject,
            body=req.body,
            cc=req.cc,
            bcc=req.bcc,
        )
    except RuntimeError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        logger.error(f"Gmail send error: {e}")
        raise HTTPException(500, f"Send failed: {e}")
    return {"status": "sent", "id": result.get("id")}


# ── AI Connectors (test) ─────────────────────────────

@app.post("/api/connectors/openai/test")
async def openai_test(current_user=Depends(get_current_user)):
    from app.connectors.openai_conn import test_connection
    try:
        return test_connection()
    except RuntimeError as e:
        raise HTTPException(400, str(e))


@app.post("/api/connectors/grok/test")
async def grok_test(current_user=Depends(get_current_user)):
    from app.connectors.grok import test_connection
    try:
        return test_connection()
    except RuntimeError as e:
        raise HTTPException(400, str(e))


@app.get("/api/ai/providers")
async def ai_providers(current_user=Depends(get_current_user)):
    """Return list of available AI providers with their status."""
    providers = []
    if settings.anthropic_api_key:
        providers.append({"id": "claude", "name": "Claude", "model": settings.claude_model, "active": True})
    if settings.grok_api_key:
        providers.append({"id": "grok", "name": "Grok (xAI)", "model": settings.grok_model, "active": True})
    if settings.openai_api_key:
        providers.append({"id": "openai", "name": "OpenAI", "model": settings.openai_model, "active": True})
    return {"providers": providers}


# ── Predictive Analytics ─────────────────────────────

@app.get("/api/analytics/predictive")
async def analytics_predictive(current_user=Depends(get_current_user)):
    try:
        return await analytics_mod.compute_analytics()
    except Exception as e:
        logger.error(f"Predictive analytics error: {e}")
        raise HTTPException(500, str(e))


# ── AI Analytics ──────────────────────────────────────

class AnalyticsRequest(BaseModel):
    prompt: str
    provider: str | None = None  # claude | grok | openai (None = auto)
    cards: list[dict] | None = None


@app.post("/api/analytics/generate")
async def analytics_generate(req: AnalyticsRequest, request: Request, current_user=Depends(get_current_user)):
    """
    AI-powered analytics: takes a natural language prompt, generates SOQL,
    runs it, then asks Claude to produce insight cards with chart configs.
    """
    if not chat_engine:
        raise HTTPException(503, "Not initialized")
    username = current_user["username"]

    from app.chat.ai_engine import _call_ai
    from app.salesforce.schema import schema_to_prompt

    schema_text = schema_to_prompt(get_schema())

    system_prompt = f"""You are a Salesforce data analyst for a staffing/consulting company.
Given the user's analytics request, generate a JSON array of analytics cards.

KEY OBJECTS & FIELDS:
- Student__c: Name, Student_Marketing_Status__c (picklist: 'In Market','Pre Marketing','Verbal Confirmation','Exit','In Job'), Technology__c, Manager__c (ref→Manager__c), Days_in_Market_Business__c, Last_Submission_Date__c, Verbal_Confirmation_Date__c, Phone__c, Email__c, Marketing_Visa_Status__c
- Submissions__c: Student_Name__c, BU_Name__c, Client_Name__c, Submission_Date__c, Offshore_Manager_Name__c, Recruiter_Name__c
- Interviews__c: Student__c (ref→Student__c), Onsite_Manager__c, Type__c, Final_Status__c (picklist: 'Good','Very Good','Average','Very Bad','Cancelled','Re-Scheduled','Confirmation','Expecting Confirmation','N/A'), Amount__c, Bill_Rate__c, Interview_Date__c
- Manager__c: Name, Active__c, Total_Expenses_MIS__c, Each_Placement_Cost__c, Students_Count__c, In_Market_Students_Count__c, Verbal_Count__c
- Job__c: Student__c, Share_With__c (ref→Manager__c), PayRate__c, Bill_Rate__c, Active__c

CROSS-OBJECT: Student__c.Manager__r.Name, Interviews__c.Student__r.Name, Job__c.Share_With__r.Name

COMMON QUERIES:
- Students in market: SELECT COUNT() FROM Student__c WHERE Student_Marketing_Status__c = 'In Market'
- Verbal confirmations: SELECT COUNT() FROM Student__c WHERE Student_Marketing_Status__c = 'Verbal Confirmation'
- Status distribution: SELECT Student_Marketing_Status__c, COUNT(Id) cnt FROM Student__c GROUP BY Student_Marketing_Status__c
- Submissions by BU: SELECT BU_Name__c, COUNT(Id) cnt FROM Submissions__c WHERE Submission_Date__c = THIS_MONTH GROUP BY BU_Name__c
- Interview status: SELECT Final_Status__c, COUNT(Id) cnt FROM Interviews__c GROUP BY Final_Status__c

Full schema:
{schema_text[:6000]}

Return ONLY valid JSON array. Each card:
{{
  "title": "Short title",
  "soql": "SELECT ... FROM ...",
  "chartType": "bar" | "pie" | "line" | "metric" | "table",
  "description": "One-line description",
  "xKey": "field for x-axis (for bar/line, use the GROUP BY field)",
  "yKey": "field for y-axis value (for bar/line, use the alias like 'cnt')",
  "labelKey": "field for pie labels (the GROUP BY field)",
  "valueKey": "field for pie values (the alias like 'cnt')"
}}

RULES:
- For "metric" cards: use SELECT COUNT() FROM ... — the totalSize in the result IS the count.
- For chart cards: use GROUP BY with COUNT(Id) aliased as cnt. Set xKey/labelKey to the grouped field and yKey/valueKey to 'cnt'.
- Use EXACT picklist values shown above (e.g. 'In Market' not 'in_market').
- Generate 3-6 cards. Only SELECT queries. No markdown.
Return ONLY the JSON array."""

    try:
        raw = await _call_ai(system_prompt, req.prompt, max_tokens=3000, provider=req.provider)
        if not raw:
            raise HTTPException(500, "No AI provider available")

        import json as _json
        text = raw.strip()
        if text.startswith("```"):
            text = text.split("\n", 1)[-1].rsplit("```", 1)[0]
        cards = _json.loads(text)

        results = []
        for card in cards:
            soql = card.get("soql", "")
            if not soql:
                continue
            try:
                result = await execute_soql(soql)
                if "error" in result:
                    card["error"] = result["error"]
                    card["records"] = []
                else:
                    records = result.get("records", [])
                    for r in records:
                        r.pop("attributes", None)
                    card["records"] = records
                    card["totalSize"] = result.get("totalSize", len(records))
            except Exception as e:
                card["error"] = str(e)
                card["records"] = []
            results.append(card)

        audit.log_action(
            username, "analytics_generate",
            {"prompt": req.prompt[:100], "cards": len(results)},
            request.client.host if request.client else "",
        )
        return {"cards": results}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Analytics generate error: {e}")
        raise HTTPException(500, str(e))


@app.post("/api/analytics/insight")
async def analytics_insight(req: AnalyticsRequest, current_user=Depends(get_current_user)):
    """Given existing card data, ask Claude to generate a written insight summary."""
    if not chat_engine:
        raise HTTPException(503, "Not initialized")

    from app.chat.ai_engine import _call_ai
    import json as _json

    cards_summary = ""
    for c in (req.cards or []):
        title = c.get("title", "")
        records = c.get("records", [])
        total = c.get("totalSize", len(records))
        cards_summary += f"\n**{title}** ({total} records): {_json.dumps(records[:10], default=str)}\n"

    system_prompt = """You are a Salesforce data analyst. Given the analytics cards and their data,
write a concise executive summary (3-5 bullet points) highlighting key insights,
trends, and actionable recommendations. Be specific with numbers. Use markdown formatting."""

    try:
        insight = await _call_ai(system_prompt, f"Analytics prompt: {req.prompt}\n\nData:\n{cards_summary}", provider=req.provider)
        if not insight:
            raise HTTPException(500, "No AI provider available")
        return {"insight": insight}
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Analytics insight error: {e}")
        raise HTTPException(500, str(e))


# ── Static (MUST be last) ─────────────────────────────

frontend_build = Path(__file__).parent.parent.parent / "frontend" / "dist"
if frontend_build.exists():
    app.mount("/", StaticFiles(directory=str(frontend_build), html=True))
